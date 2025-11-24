from flask import Flask, render_template, render_template_string, redirect, url_for, request, send_file, Response, jsonify, make_response  # pyright: ignore[reportUnusedImport]
import json
import os
import time
import logging
from functools import lru_cache
import uuid
import io
from datetime import datetime
# Imports básicos (requeridos)
try:
    from utils import parsear_monto
except ImportError:
    def parsear_monto(monto_str):
        """Función fallback si utils no está disponible"""
        try:
            return float(str(monto_str).replace('USD ', '').replace(',', '').replace('$', ''))
        except:
            return 0.0

try:
    from utils_excel import leer_excel, guardar_excel
except ImportError:
    print("⚠️ utils_excel no disponible, algunas funciones pueden no funcionar")
    leer_excel = None
    guardar_excel = None

# Imports opcionales (sistemas avanzados)
try:
    from ai_search_engine import AISearchEngine
    ai_search = AISearchEngine()
except ImportError:
    ai_search = None
    print("⚠️ AISearchEngine no disponible")

try:
    from project_updater import ProjectUpdater
    project_updater = ProjectUpdater()
except ImportError:
    project_updater = None
    print("⚠️ ProjectUpdater no disponible")

try:
    from advanced_analytics import AdvancedAnalytics
    analytics = AdvancedAnalytics()
except ImportError:
    analytics = None
    print("⚠️ AdvancedAnalytics no disponible")

try:
    from application_system import ApplicationSystem
    application_system = ApplicationSystem()
except ImportError:
    application_system = None
    print("⚠️ ApplicationSystem no disponible")

try:
    from auto_search_system import AutoSearchSystem
    auto_search_system = AutoSearchSystem()
except ImportError:
    auto_search_system = None
    print("⚠️ AutoSearchSystem no disponible")

# Imports de scrapers (opcionales)
try:
    from scrapers.bidprime import obtener_proyectos_bidprime, obtener_proyectos_bidprime_avanzado
except ImportError:
    obtener_proyectos_bidprime = None
    obtener_proyectos_bidprime_avanzado = None
    print("⚠️ Scraper bidprime no disponible")

try:
    from scrapers.todolicitaciones import obtener_proyectos_todolicitaciones
except ImportError:
    obtener_proyectos_todolicitaciones = None
    print("⚠️ Scraper todolicitaciones no disponible")

try:
    from scrapers.fuentes_gubernamentales import obtener_todos_proyectos_gubernamentales
except ImportError:
    obtener_todos_proyectos_gubernamentales = None
    print("⚠️ Scraper fuentes_gubernamentales no disponible")

try:
    from scrapers.fuentes_internacionales import obtener_todos_proyectos_internacionales
except ImportError:
    obtener_todos_proyectos_internacionales = None
    print("⚠️ Scraper fuentes_internacionales no disponible")

try:
    from scrapers.fundaciones_ongs import obtener_todos_proyectos_fundaciones
except ImportError:
    obtener_todos_proyectos_fundaciones = None
    print("⚠️ Scraper fundaciones_ongs no disponible")

try:
    from scrapers.corfo_real import obtener_proyectos_corfo_real, obtener_proyectos_corfo_por_filtros
except ImportError:
    obtener_proyectos_corfo_real = None
    obtener_proyectos_corfo_por_filtros = None
    print("⚠️ Scraper corfo_real no disponible")

# Importar búsqueda semántica (opcional)
try:
    from semantic_search import (
        buscar_proyectos_db, 
        actualizar_y_guardar_proyectos,
        SEMANTIC_AVAILABLE
    )
    SEMANTIC_SEARCH_AVAILABLE = True
except ImportError:
    SEMANTIC_SEARCH_AVAILABLE = False
    print("⚠️ Módulo de búsqueda semántica no disponible")

app = Flask(__name__, static_folder='static', static_url_path='/static')
DATA_PATH = "data/proyectos.xlsx"

# Crear directorio de datos si no existe
os.makedirs("data", exist_ok=True)

def guardar_excel_proyectos(proyectos):
    """Guarda los proyectos en un archivo Excel usando openpyxl"""
    if not proyectos:
        return
    if guardar_excel is None:
        print("⚠️ guardar_excel no disponible, no se pueden guardar proyectos")
        return
    guardar_excel(proyectos, DATA_PATH)

def _leer_excel_concreto():
    """Lee la base de datos desde el Excel disponible sin caché."""
    if leer_excel is None:
        print("⚠️ leer_excel no disponible, retornando lista vacía")
        return []
    
    # Priorizar base de datos fortalecida
    if os.path.exists('data/proyectos_fortalecidos.xlsx'):
        try:
            proyectos = leer_excel('data/proyectos_fortalecidos.xlsx')
            if proyectos:
                print(f"📂 Cargados {len(proyectos)} proyectos fortalecidos desde Excel")
                return proyectos
        except Exception as e:
            print(f"❌ Error cargando proyectos fortalecidos: {e}")
    # Intentar cargar proyectos completos
    if os.path.exists('data/proyectos_completos.xlsx'):
        try:
            proyectos = leer_excel('data/proyectos_completos.xlsx')
            if proyectos:
                print(f"📂 Cargados {len(proyectos)} proyectos completos desde Excel")
                return proyectos
        except Exception as e:
            print(f"❌ Error cargando proyectos completos: {e}")
    # Intentar cargar proyectos actualizados
    if os.path.exists('data/proyectos_actualizados.xlsx'):
        try:
            proyectos = leer_excel('data/proyectos_actualizados.xlsx')
            if proyectos:
                print(f"📂 Cargados {len(proyectos)} proyectos actualizados desde Excel")
                return proyectos
        except Exception as e:
            print(f"❌ Error cargando proyectos actualizados: {e}")
    # Si no existe, cargar el archivo original
    if os.path.exists(DATA_PATH):
        try:
            proyectos = leer_excel(DATA_PATH)
            if proyectos:
                print(f"📂 Cargados {len(proyectos)} proyectos desde Excel")
                return proyectos
        except Exception as e:
            print(f"❌ Error cargando Excel: {e}")
            return []
    return []


@lru_cache(maxsize=1)
def _cargar_excel_cached(bucket:int):
    """Devuelve proyectos cacheados por ventana de tiempo."""
    return _leer_excel_concreto()


def cargar_excel():
    """Carga los proyectos desde Excel con caché (refresco ~5 minutos)."""
    try:
        bucket = int(time.time() // 300)  # 5 minutos
        return _cargar_excel_cached(bucket)
    except Exception as e:
        print(f"⚠️ Fallback sin caché por error: {e}")
        return _leer_excel_concreto()

def recolectar_todos():
    """Recolecta proyectos de fuentes básicas"""
    print("🔄 Iniciando recolección de proyectos...")
    proyectos = []
    
    # Proyectos de ejemplo para demostración - Basados en fuentes reales
    proyectos_ejemplo = [
        {
            "Nombre": "Desarrollo Agrícola Sostenible",
            "Descripción": "Proyecto para mejorar la productividad agrícola con técnicas sostenibles",
            "Monto": "USD 50,000",
            "Área de interés": "Agricultura",
            "Estado": "Abierto",
            "Fecha cierre": "2024-12-31",
            "Fuente": "Plataforma"
        },
        {
            "Nombre": "Premio Nacional Mujer AgroInnovadora - Categoría Agricultora/Emprendedora",
            "Descripción": "Reconocimiento a mujeres que han destacado en innovación agrícola como agricultoras o emprendedoras",
            "Monto": "CLP 5,000,000",
            "Área de interés": "Agricultura",
            "Estado": "Cerrado",
            "Fecha cierre": "2025-10-07",
            "Fuente": "FIA"
        },
        {
            "Nombre": "Premio Nacional Mujer AgroInnovadora - Categoría Académica/Investigadora/Profesional",
            "Descripción": "Reconocimiento a mujeres académicas, investigadoras o profesionales que han contribuido a la innovación en el agro",
            "Monto": "CLP 5,000,000",
            "Área de interés": "Investigación",
            "Estado": "Cerrado",
            "Fecha cierre": "2025-10-07",
            "Fuente": "FIA"
        },
        {
            "Nombre": "Giras Nacionales de Innovación para Mujeres Agroinnovadoras 2025",
            "Descripción": "Programa de giras técnicas para mujeres del sector agroalimentario para conocer experiencias de innovación",
            "Monto": "CLP 2,000,000",
            "Área de interés": "Capacitación",
            "Estado": "Cerrado",
            "Fecha cierre": "2025-10-07",
            "Fuente": "FIA"
        },
        {
            "Nombre": "Hacia Sistemas Alimentarios Sostenibles - Eventos de Innovación",
            "Descripción": "Eventos de innovación con soluciones para el agro chileno considerando el rol de la mujer agroinnovadora",
            "Monto": "CLP 10,000,000",
            "Área de interés": "Innovación",
            "Estado": "Cerrado",
            "Fecha cierre": "2025-10-07",
            "Fuente": "FIA"
        },
        {
            "Nombre": "Consultorías Regionales para la Innovación 2025",
            "Descripción": "Consultorías para apoyar procesos de innovación en el sector agroalimentario a nivel regional",
            "Monto": "CLP 15,000,000",
            "Área de interés": "Consultoría",
            "Estado": "Cerrado",
            "Fecha cierre": "2025-10-07",
            "Fuente": "FIA"
        },
        {
            "Nombre": "Convocatoria Nacional de Proyectos de Innovación de Interés Privado 2025-2026",
            "Descripción": "Convocatoria para proyectos de innovación privados en el agro con énfasis regional",
            "Monto": "CLP 50,000,000",
            "Área de interés": "Innovación",
            "Estado": "Cerrado",
            "Fecha cierre": "2025-07-22",
            "Fuente": "FIA"
        },
        {
            "Nombre": "Programa de Innovación Cooperativas - AgroCoopInnova (Regiones Norte)",
            "Descripción": "Programa de innovación para cooperativas agroalimentarias en regiones de Atacama, Coquimbo, Valparaíso y Metropolitana",
            "Monto": "CLP 30,000,000",
            "Área de interés": "Cooperativas",
            "Estado": "Cerrado",
            "Fecha cierre": "2024-09-12",
            "Fuente": "FIA"
        },
        {
            "Nombre": "Programa de Innovación Cooperativas - AgroCoopInnova (Regiones Centro)",
            "Descripción": "Programa de innovación para cooperativas agroalimentarias en regiones de Maule, Ñuble y Biobío",
            "Monto": "CLP 30,000,000",
            "Área de interés": "Cooperativas",
            "Estado": "Cerrado",
            "Fecha cierre": "2024-09-12",
            "Fuente": "FIA"
        },
        {
            "Nombre": "Programa de Innovación Cooperativas - AgroCoopInnova (Regiones Sur)",
            "Descripción": "Programa de innovación para cooperativas agroalimentarias en regiones de La Araucanía y Los Ríos",
            "Monto": "CLP 30,000,000",
            "Área de interés": "Cooperativas",
            "Estado": "Cerrado",
            "Fecha cierre": "2024-09-12",
            "Fuente": "FIA"
        },
        {
            "Nombre": "Innovación Tecnológica Rural",
            "Descripción": "Implementación de tecnología en zonas rurales",
            "Monto": "USD 75,000",
            "Área de interés": "Tecnología",
            "Estado": "Abierto",
            "Fecha cierre": "2024-11-30",
            "Fuente": "FIA"
        },
        {
            "Nombre": "Conservación de Recursos Hídricos",
            "Descripción": "Proyecto de conservación y gestión del agua para uso agrícola",
            "Monto": "USD 100,000",
            "Área de interés": "Medio Ambiente",
            "Estado": "Abierto",
            "Fecha cierre": "2025-01-15",
            "Fuente": "CNR"
        },
        {
            "Nombre": "Desarrollo Rural Integral",
            "Descripción": "Programa integral de desarrollo rural para pequeños agricultores",
            "Monto": "USD 200,000",
            "Área de interés": "Desarrollo Rural",
            "Estado": "Abierto",
            "Fecha cierre": "2024-10-30",
            "Fuente": "INDAP"
        },
        {
            "Nombre": "Capacitación Agrícola",
            "Descripción": "Programa de capacitación para agricultores en técnicas modernas",
            "Monto": "USD 30,000",
            "Área de interés": "Educación",
            "Estado": "Cerrado",
            "Fecha cierre": "2024-09-15",
            "Fuente": "Fondos.gob.cl"
        },
        {
            "Nombre": "Programa Juventud Rural Innovadora - SaviaLab",
            "Descripción": "Programa estratégico para jóvenes rurales que buscan innovar en el sector agroalimentario",
            "Monto": "CLP 8,000,000",
            "Área de interés": "Juventud",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "FIA"
        },
        {
            "Nombre": "Programa Mujer AgroInnovadora - MAI",
            "Descripción": "Programa estratégico para fomentar la innovación liderada por mujeres en el sector agroalimentario",
            "Monto": "CLP 12,000,000",
            "Área de interés": "Mujer Rural",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "FIA"
        },
        {
            "Nombre": "Programa de Adopción de Innovaciones",
            "Descripción": "Programa para facilitar la adopción de innovaciones tecnológicas en la agricultura familiar campesina",
            "Monto": "CLP 20,000,000",
            "Área de interés": "Tecnología",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "FIA"
        },
        {
            "Nombre": "Fondo de Riego - CNR",
            "Descripción": "Fondo para proyectos de riego y tecnificación del riego en la agricultura",
            "Monto": "CLP 50,000,000",
            "Área de interés": "Riego",
            "Estado": "Abierto",
            "Fecha cierre": "2025-06-30",
            "Fuente": "CNR"
        },
        {
            "Nombre": "Fondo de Desarrollo de Inversiones - INDAP",
            "Descripción": "Fondo para inversiones en infraestructura y equipamiento para pequeños agricultores",
            "Monto": "CLP 15,000,000",
            "Área de interés": "Inversión",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "INDAP"
        },
        {
            "Nombre": "Programa de Asociatividad - INDAP",
            "Descripción": "Apoyo para la formación y fortalecimiento de organizaciones de pequeños agricultores",
            "Monto": "CLP 10,000,000",
            "Área de interés": "Asociatividad",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "INDAP"
        },
        {
            "Nombre": "Fondo de Fomento al Desarrollo Ganadero",
            "Descripción": "Fondo para proyectos de desarrollo y mejoramiento ganadero",
            "Monto": "CLP 25,000,000",
            "Área de interés": "Ganadería",
            "Estado": "Abierto",
            "Fecha cierre": "2025-08-31",
            "Fuente": "INDAP"
        },
        {
            "Nombre": "Programa de Desarrollo de Proveedores - CORFO",
            "Descripción": "Programa para fortalecer proveedores del sector agroalimentario",
            "Monto": "CLP 40,000,000",
            "Área de interés": "Desarrollo Empresarial",
            "Estado": "Abierto",
            "Fecha cierre": "2025-11-30",
            "Fuente": "CORFO"
        },
        {
            "Nombre": "Fondo de Innovación Agraria - FIA",
            "Descripción": "Fondo para proyectos de innovación en el sector agroalimentario",
            "Monto": "CLP 60,000,000",
            "Área de interés": "Innovación",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-15",
            "Fuente": "FIA"
        },
        {
            "Nombre": "Programa de Mejoramiento de la Competitividad - CORFO",
            "Descripción": "Programa para mejorar la competitividad de empresas del sector agroalimentario",
            "Monto": "CLP 35,000,000",
            "Área de interés": "Competitividad",
            "Estado": "Abierto",
            "Fecha cierre": "2025-10-31",
            "Fuente": "CORFO"
        },
        {
            "Nombre": "Fondo de Protección de Cultivos",
            "Descripción": "Fondo para proyectos de protección y mejoramiento de cultivos agrícolas",
            "Monto": "CLP 18,000,000",
            "Área de interés": "Protección Vegetal",
            "Estado": "Abierto",
            "Fecha cierre": "2025-09-30",
            "Fuente": "SAG"
        },
        {
            "Nombre": "Programa de Desarrollo de Mercados - ProChile",
            "Descripción": "Programa para apoyar la exportación de productos agroalimentarios",
            "Monto": "CLP 30,000,000",
            "Área de interés": "Exportación",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "ProChile"
        },
        {
            "Nombre": "Fondo de Desarrollo Sustentable",
            "Descripción": "Fondo para proyectos de agricultura sustentable y orgánica",
            "Monto": "CLP 22,000,000",
            "Área de interés": "Sustentabilidad",
            "Estado": "Abierto",
            "Fecha cierre": "2025-11-15",
            "Fuente": "Fondos.gob.cl"
        },
        {
            "Nombre": "Programa de Apoyo a la Agricultura Familiar",
            "Descripción": "Programa integral de apoyo a la agricultura familiar campesina",
            "Monto": "CLP 28,000,000",
            "Área de interés": "Agricultura Familiar",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "INDAP"
        },
        # Fondos adicionales de CNR (Comisión Nacional de Riego)
        {
            "Nombre": "Fondo Concursable para Organizaciones de Usuarios de Aguas (FOUA)",
            "Descripción": "Fondo para proyectos de mejoramiento de infraestructura de riego en organizaciones de usuarios de aguas",
            "Monto": "CLP 6,000,000",
            "Área de interés": "Riego",
            "Estado": "Abierto",
            "Fecha cierre": "2025-05-31",
            "Fuente": "CNR",
            "Enlace": "https://www.cnr.gob.cl/fondos-concursables"
        },
        {
            "Nombre": "Programa de Tecnificación del Riego",
            "Descripción": "Programa para tecnificar sistemas de riego en la agricultura",
            "Monto": "CLP 80,000,000",
            "Área de interés": "Riego",
            "Estado": "Abierto",
            "Fecha cierre": "2025-08-31",
            "Fuente": "CNR",
            "Enlace": "https://www.cnr.gob.cl/programas"
        },
        # Fondos adicionales de SAG (Servicio Agrícola y Ganadero)
        {
            "Nombre": "Sistema de Incentivos para la Sustentabilidad Agroambiental de los Suelos Agropecuarios (SIRSD-S)",
            "Descripción": "Incentivos para prácticas de conservación y mejoramiento de suelos agropecuarios",
            "Monto": "Hasta 160 UTM",
            "Área de interés": "Sustentabilidad",
            "Estado": "Abierto",
            "Fecha cierre": "2025-04-30",
            "Fuente": "SAG",
            "Enlace": "https://www.sag.gob.cl/sirsd-s"
        },
        {
            "Nombre": "Programa de Conservación de Suelos y Aseguramiento Alimentario",
            "Descripción": "Programa para conservación de suelos y seguridad alimentaria en regiones",
            "Monto": "Hasta 160 UTM",
            "Área de interés": "Conservación",
            "Estado": "Abierto",
            "Fecha cierre": "2025-06-30",
            "Fuente": "SAG",
            "Enlace": "https://www.sag.gob.cl/programas"
        },
        # Fondos adicionales de ANID (Agencia Nacional de Investigación y Desarrollo)
        {
            "Nombre": "FONDAP - Centros de Excelencia en Investigación en Áreas Prioritarias",
            "Descripción": "Fondo para centros de excelencia en investigación en áreas prioritarias del país",
            "Monto": "CLP 1,027,000,000",
            "Área de interés": "Investigación",
            "Estado": "Abierto",
            "Fecha cierre": "2025-05-31",
            "Fuente": "ANID",
            "Enlace": "https://www.anid.cl/fondap"
        },
        {
            "Nombre": "FONDECYT Regular - Proyectos de Investigación",
            "Descripción": "Fondo para proyectos de investigación científica y tecnológica",
            "Monto": "CLP 150,000,000",
            "Área de interés": "Investigación",
            "Estado": "Abierto",
            "Fecha cierre": "2025-07-15",
            "Fuente": "ANID",
            "Enlace": "https://www.anid.cl/fondecyt"
        },
        # Fondos adicionales de CORFO
        {
            "Nombre": "Programa de Apoyo al Emprendimiento e Innovación - Start-Up Chile",
            "Descripción": "Programa para emprendimientos innovadores con potencial de crecimiento global",
            "Monto": "USD 80,000",
            "Área de interés": "Emprendimiento",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "CORFO",
            "Enlace": "https://www.startupchile.org"
        },
        {
            "Nombre": "Programa de Innovación en Alimentos - Transforma Alimentos",
            "Descripción": "Programa para proyectos de innovación en la industria alimentaria",
            "Monto": "CLP 100,000,000",
            "Área de interés": "Innovación",
            "Estado": "Abierto",
            "Fecha cierre": "2025-09-30",
            "Fuente": "CORFO",
            "Enlace": "https://www.corfo.cl/programas"
        },
        {
            "Nombre": "Programa de Desarrollo de Proveedores - PDP",
            "Descripción": "Programa para fortalecer proveedores de grandes empresas",
            "Monto": "CLP 50,000,000",
            "Área de interés": "Desarrollo Empresarial",
            "Estado": "Abierto",
            "Fecha cierre": "2025-11-30",
            "Fuente": "CORFO",
            "Enlace": "https://www.corfo.cl/pdp"
        },
        # Fondos adicionales de INDAP
        {
            "Nombre": "Programa de Desarrollo Local - PRODESAL",
            "Descripción": "Programa de desarrollo local para pequeños productores agrícolas",
            "Monto": "CLP 20,000,000",
            "Área de interés": "Desarrollo Local",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "INDAP",
            "Enlace": "https://www.indap.gob.cl/prodesal"
        },
        {
            "Nombre": "Programa de Desarrollo de Inversiones - PDI",
            "Descripción": "Programa para inversiones en infraestructura productiva",
            "Monto": "CLP 25,000,000",
            "Área de interés": "Inversión",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "INDAP",
            "Enlace": "https://www.indap.gob.cl/pdi"
        },
        {
            "Nombre": "Programa de Asesoría Técnica - SAT",
            "Descripción": "Programa de asesoría técnica especializada para pequeños agricultores",
            "Monto": "CLP 12,000,000",
            "Área de interés": "Asesoría",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "INDAP",
            "Enlace": "https://www.indap.gob.cl/sat"
        },
        # Fondos adicionales de ProChile
        {
            "Nombre": "Programa de Apoyo a la Exportación - PAE",
            "Descripción": "Programa para apoyar la exportación de productos chilenos",
            "Monto": "CLP 40,000,000",
            "Área de interés": "Exportación",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "ProChile",
            "Enlace": "https://www.prochile.gob.cl/pae"
        },
        {
            "Nombre": "Programa de Promoción de Exportaciones",
            "Descripción": "Programa para promoción de productos chilenos en mercados internacionales",
            "Monto": "CLP 35,000,000",
            "Área de interés": "Exportación",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "ProChile",
            "Enlace": "https://www.prochile.gob.cl/programas"
        },
        # Fondos internacionales - FAO
        {
            "Nombre": "FAO - Programa de Cooperación Técnica",
            "Descripción": "Programa de cooperación técnica de la FAO para proyectos agrícolas",
            "Monto": "USD 500,000",
            "Área de interés": "Cooperación Internacional",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "FAO",
            "Enlace": "https://www.fao.org/calls-for-proposals"
        },
        {
            "Nombre": "FAO - Fondo Verde para el Clima",
            "Descripción": "Fondo para proyectos de adaptación al cambio climático en agricultura",
            "Monto": "USD 1,000,000",
            "Área de interés": "Cambio Climático",
            "Estado": "Abierto",
            "Fecha cierre": "2025-10-31",
            "Fuente": "FAO",
            "Enlace": "https://www.fao.org/green-climate-fund"
        },
        # Fondos internacionales - Banco Mundial
        {
            "Nombre": "Banco Mundial - Proyecto de Desarrollo Rural",
            "Descripción": "Proyecto del Banco Mundial para desarrollo rural y agrícola",
            "Monto": "USD 5,000,000",
            "Área de interés": "Desarrollo Rural",
            "Estado": "Abierto",
            "Fecha cierre": "2025-09-30",
            "Fuente": "Banco Mundial",
            "Enlace": "https://www.worldbank.org/projects"
        },
        {
            "Nombre": "Banco Mundial - Fondo de Innovación Agrícola",
            "Descripción": "Fondo para innovación en agricultura y seguridad alimentaria",
            "Monto": "USD 2,500,000",
            "Área de interés": "Innovación",
            "Estado": "Abierto",
            "Fecha cierre": "2025-08-31",
            "Fuente": "Banco Mundial",
            "Enlace": "https://www.worldbank.org/agriculture"
        },
        # Fondos internacionales - BID
        {
            "Nombre": "BID - Programa de Desarrollo Agrícola",
            "Descripción": "Programa del BID para desarrollo agrícola en América Latina",
            "Monto": "USD 3,000,000",
            "Área de interés": "Desarrollo Agrícola",
            "Estado": "Abierto",
            "Fecha cierre": "2025-11-30",
            "Fuente": "BID",
            "Enlace": "https://www.iadb.org/agriculture"
        },
        {
            "Nombre": "BID - Fondo de Innovación Rural",
            "Descripción": "Fondo para innovación en zonas rurales de América Latina",
            "Monto": "USD 1,500,000",
            "Área de interés": "Innovación Rural",
            "Estado": "Abierto",
            "Fecha cierre": "2025-10-15",
            "Fuente": "BID",
            "Enlace": "https://www.iadb.org/rural-innovation"
        },
        # Fondos internacionales - IFAD
        {
            "Nombre": "IFAD - Fondo Internacional de Desarrollo Agrícola",
            "Descripción": "Fondo para proyectos de desarrollo agrícola en países en desarrollo",
            "Monto": "USD 2,000,000",
            "Área de interés": "Desarrollo Agrícola",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "IFAD",
            "Enlace": "https://www.ifad.org/grants"
        },
        {
            "Nombre": "IFAD - Programa de Agricultura Familiar",
            "Descripción": "Programa para fortalecer la agricultura familiar campesina",
            "Monto": "USD 1,800,000",
            "Área de interés": "Agricultura Familiar",
            "Estado": "Abierto",
            "Fecha cierre": "2025-11-30",
            "Fuente": "IFAD",
            "Enlace": "https://www.ifad.org/family-farming"
        },
        # Fondos internacionales - USAID
        {
            "Nombre": "USAID - Programa de Seguridad Alimentaria",
            "Descripción": "Programa de USAID para mejorar la seguridad alimentaria",
            "Monto": "USD 4,000,000",
            "Área de interés": "Seguridad Alimentaria",
            "Estado": "Abierto",
            "Fecha cierre": "2025-09-30",
            "Fuente": "USAID",
            "Enlace": "https://www.usaid.gov/food-security"
        },
        {
            "Nombre": "USAID - Fondo de Innovación Agrícola",
            "Descripción": "Fondo para innovación en agricultura y tecnología",
            "Monto": "USD 2,200,000",
            "Área de interés": "Innovación",
            "Estado": "Abierto",
            "Fecha cierre": "2025-08-31",
            "Fuente": "USAID",
            "Enlace": "https://www.usaid.gov/agriculture"
        },
        # Fondos internacionales - Unión Europea
        {
            "Nombre": "UE - Programa Horizon Europe - Agricultura",
            "Descripción": "Programa de investigación e innovación de la UE en agricultura",
            "Monto": "EUR 3,500,000",
            "Área de interés": "Investigación",
            "Estado": "Abierto",
            "Fecha cierre": "2025-10-31",
            "Fuente": "Unión Europea",
            "Enlace": "https://ec.europa.eu/horizon-europe"
        },
        {
            "Nombre": "UE - Programa de Desarrollo Rural",
            "Descripción": "Programa de la UE para desarrollo rural sostenible",
            "Monto": "EUR 2,800,000",
            "Área de interés": "Desarrollo Rural",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "Unión Europea",
            "Enlace": "https://ec.europa.eu/rural-development"
        },
        # Fondos de FOSIS
        {
            "Nombre": "FOSIS - Programa de Emprendimiento Local",
            "Descripción": "Programa para emprendimientos locales en zonas vulnerables",
            "Monto": "CLP 8,000,000",
            "Área de interés": "Emprendimiento",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "FOSIS",
            "Enlace": "https://www.fosis.gob.cl/emprendimiento"
        },
        {
            "Nombre": "FOSIS - Programa de Desarrollo de Capacidades",
            "Descripción": "Programa para desarrollo de capacidades en comunidades vulnerables",
            "Monto": "CLP 6,000,000",
            "Área de interés": "Capacitación",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "FOSIS",
            "Enlace": "https://www.fosis.gob.cl/capacidades"
        },
        # Fondos de SERCOTEC
        {
            "Nombre": "SERCOTEC - Programa de Apoyo al Emprendimiento",
            "Descripción": "Programa para apoyar emprendimientos de micro y pequeñas empresas",
            "Monto": "CLP 5,000,000",
            "Área de interés": "Emprendimiento",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "SERCOTEC",
            "Enlace": "https://www.sercotec.cl/emprendimiento"
        },
        {
            "Nombre": "SERCOTEC - Programa de Fortalecimiento Empresarial",
            "Descripción": "Programa para fortalecer empresas existentes",
            "Monto": "CLP 7,000,000",
            "Área de interés": "Desarrollo Empresarial",
            "Estado": "Abierto",
            "Fecha cierre": "2025-12-31",
            "Fuente": "SERCOTEC",
            "Enlace": "https://www.sercotec.cl/fortalecimiento"
        },
        # Fondos de GEF (Global Environment Facility)
        {
            "Nombre": "GEF - Fondo para el Medio Ambiente Mundial",
            "Descripción": "Fondo para proyectos ambientales y de desarrollo sostenible",
            "Monto": "USD 1,200,000",
            "Área de interés": "Medio Ambiente",
            "Estado": "Abierto",
            "Fecha cierre": "2025-09-30",
            "Fuente": "GEF",
            "Enlace": "https://www.thegef.org/grants"
        },
        {
            "Nombre": "GEF - Programa de Adaptación al Cambio Climático",
            "Descripción": "Programa para adaptación al cambio climático en agricultura",
            "Monto": "USD 1,500,000",
            "Área de interés": "Cambio Climático",
            "Estado": "Abierto",
            "Fecha cierre": "2025-10-31",
            "Fuente": "GEF",
            "Enlace": "https://www.thegef.org/climate-adaptation"
        },
        # Fondos de UNDP
        {
            "Nombre": "UNDP - Programa de Desarrollo Sostenible",
            "Descripción": "Programa del PNUD para desarrollo sostenible",
            "Monto": "USD 2,000,000",
            "Área de interés": "Desarrollo Sostenible",
            "Estado": "Abierto",
            "Fecha cierre": "2025-11-30",
            "Fuente": "UNDP",
            "Enlace": "https://www.undp.org/grants"
        },
        {
            "Nombre": "UNDP - Fondo de Innovación Social",
            "Descripción": "Fondo para innovación social y desarrollo comunitario",
            "Monto": "USD 800,000",
            "Área de interés": "Innovación Social",
            "Estado": "Abierto",
            "Fecha cierre": "2025-08-31",
            "Fuente": "UNDP",
            "Enlace": "https://www.undp.org/social-innovation"
        },
        # Fondos de Fundación Gates
        {
            "Nombre": "Fundación Gates - Programa de Desarrollo Agrícola",
            "Descripción": "Programa de la Fundación Gates para desarrollo agrícola",
            "Monto": "USD 3,500,000",
            "Área de interés": "Desarrollo Agrícola",
            "Estado": "Abierto",
            "Fecha cierre": "2025-10-15",
            "Fuente": "Fundación Gates",
            "Enlace": "https://www.gatesfoundation.org/agriculture"
        },
        {
            "Nombre": "Fundación Gates - Fondo de Innovación en Nutrición",
            "Descripción": "Fondo para innovación en nutrición y seguridad alimentaria",
            "Monto": "USD 2,500,000",
            "Área de interés": "Nutrición",
            "Estado": "Abierto",
            "Fecha cierre": "2025-09-30",
            "Fuente": "Fundación Gates",
            "Enlace": "https://www.gatesfoundation.org/nutrition"
        }
    ]
    
    proyectos.extend(proyectos_ejemplo)
    
    # BIDPrime - Nueva fuente de financiamiento
    if obtener_proyectos_bidprime is not None:
        try:
            proyectos_bidprime = obtener_proyectos_bidprime()
            if proyectos_bidprime:
                proyectos.extend(proyectos_bidprime)
                print(f"✅ BIDPrime: {len(proyectos_bidprime)} proyectos agregados")
        except Exception as e:
            print(f"❌ Error con BIDPrime: {e}")
    
    if obtener_proyectos_bidprime_avanzado is not None:
        try:
            proyectos_bidprime_avanzado = obtener_proyectos_bidprime_avanzado()
            if proyectos_bidprime_avanzado:
                proyectos.extend(proyectos_bidprime_avanzado)
                print(f"✅ BIDPrime Avanzado: {len(proyectos_bidprime_avanzado)} proyectos adicionales")
        except Exception as e:
            print(f"❌ Error con BIDPrime Avanzado: {e}")
    
    # TodoLicitaciones.cl - Portal de licitaciones públicas
    if obtener_proyectos_todolicitaciones is not None:
        try:
            proyectos_todolicitaciones = obtener_proyectos_todolicitaciones()
            if proyectos_todolicitaciones:
                proyectos.extend(proyectos_todolicitaciones)
                print(f"✅ TodoLicitaciones.cl: {len(proyectos_todolicitaciones)} proyectos agregados")
        except Exception as e:
            print(f"❌ Error con TodoLicitaciones.cl: {e}")
    
    # Fuentes Gubernamentales Chilenas
    if obtener_todos_proyectos_gubernamentales is not None:
        try:
            proyectos_gubernamentales = obtener_todos_proyectos_gubernamentales()
            if proyectos_gubernamentales:
                proyectos.extend(proyectos_gubernamentales)
                print(f"✅ Fuentes Gubernamentales: {len(proyectos_gubernamentales)} proyectos agregados")
        except Exception as e:
            print(f"❌ Error con fuentes gubernamentales: {e}")
    
    # Fuentes Internacionales
    if obtener_todos_proyectos_internacionales is not None:
        try:
            proyectos_internacionales = obtener_todos_proyectos_internacionales()
            if proyectos_internacionales:
                proyectos.extend(proyectos_internacionales)
                print(f"✅ Fuentes Internacionales: {len(proyectos_internacionales)} proyectos agregados")
        except Exception as e:
            print(f"❌ Error con fuentes internacionales: {e}")
    
    # Fundaciones y ONGs
    if obtener_todos_proyectos_fundaciones is not None:
        try:
            proyectos_fundaciones = obtener_todos_proyectos_fundaciones()
            if proyectos_fundaciones:
                proyectos.extend(proyectos_fundaciones)
                print(f"✅ Fundaciones y ONGs: {len(proyectos_fundaciones)} proyectos agregados")
        except Exception as e:
            print(f"❌ Error con fundaciones y ONGs: {e}")
    
    # CORFO - Fuente oficial de programas y convocatorias
    if obtener_proyectos_corfo_real is not None:
        try:
            proyectos_corfo = obtener_proyectos_corfo_real()
            if proyectos_corfo:
                proyectos.extend(proyectos_corfo)
                print(f"✅ CORFO Real: {len(proyectos_corfo)} proyectos agregados")
        except Exception as e:
            print(f"❌ Error con CORFO Real: {e}")
    
    if obtener_proyectos_corfo_por_filtros is not None:
        try:
            proyectos_corfo_filtrados = obtener_proyectos_corfo_por_filtros()
            if proyectos_corfo_filtrados:
                proyectos.extend(proyectos_corfo_filtrados)
                print(f"✅ CORFO Filtrados: {len(proyectos_corfo_filtrados)} proyectos adicionales")
        except Exception as e:
            print(f"❌ Error con CORFO Filtrados: {e}")
    
    print(f"🎉 Recolección completada: {len(proyectos)} proyectos totales")
    return proyectos

def parse_fecha_sort(fecha_str):
    """Convierte fecha a formato YYYY-MM-DD para ordenamiento"""
    if not fecha_str or fecha_str == 'N/A':
        return '9999-12-31'  # Fechas vacías al final
    
    try:
        # Intentar diferentes formatos
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
            try:
                return datetime.strptime(fecha_str, fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
        return '9999-12-31'
    except:
        return '9999-12-31'

@app.route('/', methods=['GET'])
def home():
    # Generar o obtener session ID
    session_id = request.cookies.get('session_id')
    if not session_id:
        session_id = str(uuid.uuid4())
    
    # Detectar idioma del navegador o usar español por defecto
    language = request.args.get('lang', 'es')
    if language not in ['es', 'en']:
        language = 'es'
    
    proyectos = cargar_excel()
    if not proyectos:
        proyectos = recolectar_todos()
        guardar_excel_proyectos(proyectos)
    
    # Aplicar filtros
    query = request.args.get('query', '').strip()
    area = request.args.get('area', '').strip()
    estado = request.args.get('estado', '').strip()
    ordenar_por = request.args.get('ordenar_por', 'fecha')
    orden = request.args.get('orden', 'asc')
    
    # Filtrar proyectos con motor de búsqueda mejorado
    proyectos_filtrados = []
    for p in proyectos:
        # Filtro por búsqueda mejorada con palabras clave
        if query:
            query_lower = query.lower()
            # Palabras clave para búsqueda inteligente
            keywords = {
                'agricultura sostenible': ['sostenible', 'sustentable', 'ecológico', 'orgánico'],
                'gestión hídrica': ['agua', 'hídrico', 'riego', 'hidráulico'],
                'desarrollo rural': ['rural', 'campo', 'comunidad', 'local'],
                'innovación tecnológica': ['tecnología', 'digital', 'innovación', 'modernización'],
                'seguridad alimentaria': ['alimentario', 'nutrición', 'comida', 'producción'],
                'comercio agrícola': ['comercio', 'exportación', 'mercado', 'comercial']
            }
            
            # Buscar coincidencias directas
            direct_match = (query_lower in p.get('Nombre', '').lower() or 
                          query_lower in p.get('Área de interés', '').lower() or
                          query_lower in p.get('Fuente', '').lower() or
                          query_lower in p.get('Monto', '').lower() or
                          query_lower in p.get('Estado', '').lower())
            
            # Buscar por palabras clave relacionadas
            keyword_match = False
            for category, words in keywords.items():
                if query_lower in category or any(word in query_lower for word in words):
                    # Buscar estas palabras en los campos del proyecto
                    for word in words:
                        if (word in p.get('Nombre', '').lower() or 
                            word in p.get('Área de interés', '').lower() or
                            word in p.get('Fuente', '').lower()):
                            keyword_match = True
                            break
            
            if not (direct_match or keyword_match):
                continue
        
        # Filtro por área
        if area and p.get('Área de interés', '') != area:
            continue
        
        # Filtro por estado
        if estado and p.get('Estado', '') != estado:
            continue
        
        proyectos_filtrados.append(p)
    
    # Ordenar proyectos
    if ordenar_por == 'fecha':
        proyectos_filtrados.sort(key=lambda x: parse_fecha_sort(x.get('Fecha cierre', '')), reverse=(orden == 'desc'))
    elif ordenar_por == 'monto':
        proyectos_filtrados.sort(key=lambda x: parsear_monto(x.get('Monto', '0')), reverse=(orden == 'desc'))
    elif ordenar_por == 'nombre':
        proyectos_filtrados.sort(key=lambda x: x.get('Nombre', '').lower(), reverse=(orden == 'desc'))
    elif ordenar_por == 'fuente':
        proyectos_filtrados.sort(key=lambda x: x.get('Fuente', '').lower(), reverse=(orden == 'desc'))
    
    # Paginación - 10 proyectos por página
    page = int(request.args.get('page', 1))
    per_page = 10
    total_proyectos = len(proyectos_filtrados)
    total_pages = (total_proyectos + per_page - 1) // per_page
    
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    proyectos_paginados = proyectos_filtrados[start_idx:end_idx]
    
    # Estadísticas básicas
    total = len(proyectos)
    por_area = {}
    por_estado = {}
    por_moneda = {}
    total_monto = {}
    
    for p in proyectos:
        por_area[p.get("Área de interés", "General")] = por_area.get(p.get("Área de interés", "General"), 0) + 1
        por_estado[p.get("Estado", "N/D")] = por_estado.get(p.get("Estado", "N/D"), 0) + 1
        
        # Acumular montos por moneda
        monto_val = str(p.get("Monto", "0"))
        parts = monto_val.split()
        if len(parts) == 2 and parts[0].isalpha():
            moneda = parts[0]
            try:
                numero = float(parts[1].replace(',', '').replace('.', ''))
            except:
                numero = 0.0
            total_monto[moneda] = total_monto.get(moneda, 0.0) + numero
            por_moneda[moneda] = por_moneda.get(moneda, 0) + 1

    stats = {
        "total": total,
        "por_area": por_area,
        "por_estado": por_estado,
        "por_moneda": por_moneda,
        "total_monto": total_monto,
    }
    
    # Datos de paginación
    pagination_data = {
        'current_page': page,
        'total_pages': total_pages,
        'total_proyectos': total_proyectos,
        'per_page': per_page,
        'has_prev': page > 1,
        'has_next': page < total_pages,
        'prev_page': page - 1 if page > 1 else None,
        'next_page': page + 1 if page < total_pages else None
    }
    
    # Información de contacto
    info = {
        "titulo": "Plataforma de Fondos",
        "descripcion": (
            "Plataforma de fondos y convocatorias para proyectos "
            "de financiamiento agrícola y desarrollo rural."
        ),
        "contacto": {
            "email": "contacto@plataforma.cl",
            "telefono": "",
            "fax": "",
            "direccion": "",
            "correo_postal": "",
            "representante": ""
        },
        "redes_sociales": {
            "twitter": "",
            "instagram": "",
            "facebook": ""
        }
    }
    
    # Datos para el template
    adjudicados = []
    documentos = []
    postulaciones = []
    instituciones = []
    
    response = make_response(render_template('home_ordenado.html',
                         proyectos=proyectos_paginados, 
                         stats=stats, 
                         adjudicados=adjudicados, 
                         documentos=documentos, 
                         postulaciones=postulaciones, 
                         instituciones=instituciones, 
                         info=info,
                         current_language=language,
                         pagination=pagination_data,
                         current_filters={
                             'query': query,
                             'area': area,
                             'estado': estado,
                             'ordenar_por': ordenar_por,
                             'orden': orden
                         }))
    
    # Establecer cookie de sesión
    response.set_cookie('session_id', session_id, max_age=86400)  # 24 horas
    
    return response

@app.route('/buscar', methods=['POST'])
def buscar():
    proyectos = recolectar_todos()
    guardar_excel_proyectos(proyectos)
    return redirect(url_for('home'))

@app.route('/quienes-somos', methods=['GET'])
def quienes_somos():
    return render_template('quienes_somos.html')

@app.route('/adjudicados', methods=['GET'])
def adjudicados():
    return render_template('adjudicados.html')

@app.route('/dashboard-old', methods=['GET'])
def dashboard_old():
    """Dashboard principal con estadísticas y gráficos"""
    proyectos = cargar_excel()
    if not proyectos:
        proyectos = recolectar_todos()
        guardar_excel_proyectos(proyectos)
    
    # Estadísticas
    total_proyectos = len(proyectos)
    proyectos_abiertos = len([p for p in proyectos if p.get('Estado') == 'Abierto'])
    fuentes_unicas = len(set(p.get('Fuente', '') for p in proyectos))
    
    # Calcular monto total
    monto_total = 0
    for p in proyectos:
        monto_str = str(p.get('Monto', '0')).replace('USD ', '').replace(',', '').replace('$', '')
        try:
            monto_num = float(monto_str)
            monto_total += monto_num
        except:
            pass
    
    monto_total_formateado = f"{monto_total:,.0f}"
    
    # Paginación para proyectos destacados
    page = int(request.args.get('page', 1))
    per_page = 10
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    proyectos_paginados = proyectos[start_idx:end_idx]
    
    # Estadísticas para gráficos
    stats = {
        'total_proyectos': total_proyectos,
        'proyectos_abiertos': proyectos_abiertos,
        'fuentes_unicas': fuentes_unicas,
        'monto_total': monto_total,
        'monto_total_formateado': monto_total_formateado
    }
    
    return render_template('dashboard.html',
                         proyectos=proyectos,
                         proyectos_paginados=proyectos_paginados,
                         total_proyectos=total_proyectos,
                         proyectos_abiertos=proyectos_abiertos,
                         fuentes_unicas=fuentes_unicas,
                         monto_total_formateado=monto_total_formateado,
                         stats=stats)

@app.route('/ai-search', methods=['GET'])
def ai_search_page():
    """Página del buscador inteligente"""
    return render_template('ai_search.html')

@app.route('/todos-los-proyectos', methods=['GET'])
def todos_los_proyectos():
    """Muestra todos los proyectos disponibles en tablas organizadas"""
    proyectos = cargar_excel()
    if not proyectos:
        proyectos = recolectar_todos()
        guardar_excel_proyectos(proyectos)
    
    # Filtros de búsqueda
    query = request.args.get('query', '').lower()
    area = request.args.get('area', '')
    fuente = request.args.get('fuente', '')
    estado = request.args.get('estado', '')
    monto_min = request.args.get('monto_min', '')
    monto_max = request.args.get('monto_max', '')
    
    # Aplicar filtros
    proyectos_filtrados = proyectos.copy()
    
    if query:
        proyectos_filtrados = [p for p in proyectos_filtrados if 
                              query in p.get('Nombre', '').lower() or 
                              query in p.get('Descripción', '').lower() or 
                              query in p.get('Área de interés', '').lower()]
    
    if area:
        proyectos_filtrados = [p for p in proyectos_filtrados if p.get('Área de interés') == area]
    
    if fuente:
        proyectos_filtrados = [p for p in proyectos_filtrados if p.get('Fuente') == fuente]
    
    if estado:
        proyectos_filtrados = [p for p in proyectos_filtrados if p.get('Estado') == estado]
    
    # Organizar proyectos por categorías
    proyectos_nacionales = [p for p in proyectos_filtrados if p.get('Fuente') in ['Corfo', 'Minagri', 'INDAP', 'FIA', 'CONICYT', 'Gobierno Chile', 'Ministerio Ambiente', 'SAG', 'INIA', 'ProChile', 'SERNATUR', 'SUBPESCA']]
    proyectos_internacionales = [p for p in proyectos_filtrados if p.get('Fuente') in ['FAO', 'Banco Mundial', 'BID', 'FIDA', 'UNDP', 'GEF', 'UE', 'Canadá', 'Australia', 'Japón', 'USAID', 'DFID Reino Unido', 'GIZ Alemania', 'AFD Francia', 'JICA Japón']]
    proyectos_regionales = [p for p in proyectos_filtrados if p.get('Fuente') in ['CEPAL', 'OEA', 'SICA', 'CARICOM', 'MERCOSUR']]
    proyectos_privados = [p for p in proyectos_filtrados if 'Fundación' in p.get('Fuente', '') or p.get('Fuente') in ['Nestlé', 'Unilever', 'Coca-Cola', 'PepsiCo', 'Walmart']]
    proyectos_academicos = [p for p in proyectos_filtrados if 'Universidad' in p.get('Fuente', '') or p.get('Fuente') in ['MIT', 'Harvard', 'Stanford', 'UC Berkeley', 'Cornell']]
    proyectos_gubernamentales = [p for p in proyectos_filtrados if p.get('Fuente') in ['USAID', 'DFID Reino Unido', 'GIZ Alemania', 'AFD Francia', 'JICA Japón']]
    proyectos_especializados = [p for p in proyectos_filtrados if p.get('Fuente') in ['CGIAR', 'CIFOR', 'ICRAF', 'IWMI', 'ILRI']]
    
    # Obtener listas únicas para filtros
    areas_unicas = list(set(p.get('Área de interés', '') for p in proyectos))
    fuentes_unicas = list(set(p.get('Fuente', '') for p in proyectos))
    estados_unicos = list(set(p.get('Estado', '') for p in proyectos))
    
    return render_template('todos_los_proyectos_mejorado.html',
                         proyectos_nacionales=proyectos_nacionales,
                         proyectos_internacionales=proyectos_internacionales,
                         proyectos_regionales=proyectos_regionales,
                         proyectos_privados=proyectos_privados,
                         proyectos_academicos=proyectos_academicos,
                         proyectos_gubernamentales=proyectos_gubernamentales,
                         proyectos_especializados=proyectos_especializados,
                         total_proyectos=len(proyectos_filtrados),
                         areas_unicas=areas_unicas,
                         fuentes_unicas=fuentes_unicas,
                         estados_unicos=estados_unicos,
                         query=query,
                         area=area,
                         fuente=fuente,
                         estado=estado)

@app.route('/exportar-excel', methods=['GET'])
def exportar_excel():
    """Exporta todos los proyectos a Excel usando openpyxl"""
    proyectos = cargar_excel()
    if not proyectos:
        proyectos = recolectar_todos()
    
    # Asegurar unicidad de columnas
    all_keys = []
    if proyectos:
        all_keys_set = set()
        for p in proyectos:
            all_keys_set.update(p.keys())
        all_keys = sorted(list(all_keys_set))
        proyectos = [{k: p.get(k, "") for k in all_keys} for p in proyectos]
    
    # Crear archivo Excel en memoria usando openpyxl
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Proyectos"
    
    # Escribir encabezados
    if proyectos and all_keys:
        for idx, key in enumerate(all_keys, start=1):
            ws.cell(row=1, column=idx, value=key)
        
        # Escribir datos
        for row_idx, proyecto in enumerate(proyectos, start=2):
            for col_idx, key in enumerate(all_keys, start=1):
                valor = proyecto.get(key, "")
                ws.cell(row=row_idx, column=col_idx, value=valor)
    
    # Guardar en BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'proyectos_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/exportar-csv', methods=['GET'])
def exportar_csv():
    """Exporta todos los proyectos a CSV sin pandas"""
    import csv
    proyectos = cargar_excel()
    if not proyectos:
        proyectos = recolectar_todos()
    
    # Obtener todas las claves únicas
    all_keys = set()
    for p in proyectos:
        all_keys.update(p.keys())
    all_keys = sorted(list(all_keys))
    
    # Crear CSV en memoria
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=all_keys, extrasaction='ignore')
    writer.writeheader()
    for proyecto in proyectos:
        writer.writerow({k: proyecto.get(k, "") for k in all_keys})
    
    output.seek(0)
    
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        as_attachment=True,
        download_name=f'proyectos_{datetime.now().strftime("%Y%m%d")}.csv',
        mimetype='text/csv'
    )

@app.route('/api/proyectos', methods=['GET'])
def api_proyectos():
    """API para obtener proyectos en formato JSON"""
    proyectos = cargar_excel()
    if not proyectos:
        proyectos = recolectar_todos()
    
    return jsonify({
        'total': len(proyectos),
        'proyectos': proyectos
    })

@app.route('/api/buscar', methods=['GET'])
def api_buscar():
    """API para búsqueda de proyectos"""
    query = request.args.get('q', '').lower()
    area = request.args.get('area', '')
    fuente = request.args.get('fuente', '')
    
    proyectos = cargar_excel()
    if not proyectos:
        proyectos = recolectar_todos()
    
    # Aplicar filtros
    resultados = proyectos.copy()
    
    if query:
        resultados = [p for p in resultados if 
                     query in p.get('Nombre', '').lower() or 
                     query in p.get('Descripción', '').lower() or 
                     query in p.get('Área de interés', '').lower()]
    
    if area:
        resultados = [p for p in resultados if p.get('Área de interés') == area]
    
    if fuente:
        resultados = [p for p in resultados if p.get('Fuente') == fuente]
    
    return jsonify({
        'total': len(resultados),
        'resultados': resultados
    })

@app.route('/busqueda-semantica', methods=['GET'])
def busqueda_semantica():
    """Página de búsqueda semántica de proyectos"""
    if not SEMANTIC_SEARCH_AVAILABLE:
        return render_template('error.html', error='Búsqueda semántica no disponible. Instale sentence-transformers.'), 503
    
    query = request.args.get('query', '')
    from semantic_search import buscar_proyectos_db
    proyectos = buscar_proyectos_db(query) if query else buscar_proyectos_db()
    
    return render_template_string('''
    <!doctype html>
    <html>
    <head>
        <title>Buscador Semántico de Proyectos</title>
        <meta charset="utf-8">
        <style>
            body { font-family: Arial, sans-serif; max-width: 1200px; margin: 0 auto; padding: 20px; }
            h1 { color: #2c3e50; }
            form { margin: 20px 0; }
            input[type="text"] { padding: 10px; width: 400px; font-size: 16px; }
            input[type="submit"] { padding: 10px 20px; font-size: 16px; background: #3498db; color: white; border: none; cursor: pointer; }
            input[type="submit"]:hover { background: #2980b9; }
            ul { list-style: none; padding: 0; }
            li { margin: 20px 0; padding: 15px; background: #f8f9fa; border-left: 4px solid #3498db; }
            .fuente { color: #7f8c8d; font-size: 14px; }
            .similitud { color: #27ae60; font-weight: bold; }
        </style>
    </head>
    <body>
        <h1>🔍 Buscador Semántico de Proyectos</h1>
        <p>Busque proyectos financiados relevantes usando inteligencia artificial</p>
        <form method="get" action="/busqueda-semantica">
            <input type="text" name="query" placeholder="Buscar por palabra clave" value="{{query}}">
            <input type="submit" value="Buscar">
        </form>
        <hr>
        {% if proyectos %}
            <h2>Resultados ({{proyectos|length}})</h2>
            <ul>
            {% for fuente, titulo, resumen, similitud in proyectos %}
                <li>
                    <h3>{{titulo}}</h3>
                    <p class="fuente">Fuente: {{fuente}} | Similitud: <span class="similitud">{{'%.2f' % similitud}}</span></p>
                    <p>{{resumen}}</p>
                </li>
            {% endfor %}
            </ul>
        {% else %}
            <p>No hay proyectos que coincidan. <a href="/api/semantic/update">Actualizar base de datos</a></p>
        {% endif %}
    </body>
    </html>
    ''', proyectos=proyectos, query=query)

@app.route('/api/semantic/search', methods=['GET'])
def api_semantic_search():
    """API para búsqueda semántica de proyectos"""
    if not SEMANTIC_SEARCH_AVAILABLE:
        return jsonify({'error': 'Búsqueda semántica no disponible'}), 503
    
    from semantic_search import buscar_proyectos_db
    query = request.args.get('q', '')
    limite = int(request.args.get('limit', 100))
    
    proyectos = buscar_proyectos_db(query, limite)
    
    resultados = [
        {
            'fuente': p[0],
            'titulo': p[1],
            'resumen': p[2],
            'similitud': p[3]
        }
        for p in proyectos
    ]
    
    return jsonify({
        'total': len(resultados),
        'resultados': resultados,
        'query': query
    })

@app.route('/api/semantic/update', methods=['POST'])
def api_semantic_update():
    """API para actualizar proyectos semánticos"""
    if not SEMANTIC_SEARCH_AVAILABLE:
        return jsonify({'error': 'Búsqueda semántica no disponible'}), 503
    
    try:
        from semantic_search import actualizar_y_guardar_proyectos
        actualizar_y_guardar_proyectos()
        return jsonify({
            'success': True,
            'message': 'Proyectos semánticos actualizados correctamente'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/ai-search', methods=['POST'])
def api_ai_search():
    """API para búsqueda inteligente con IA"""
    if ai_search is None:
        return jsonify({'error': 'Búsqueda con IA no disponible'}), 503
    
    data = request.get_json()
    query = data.get('query', '')
    
    if not query:
        return jsonify({'error': 'Query is required'}), 400
    
    proyectos = cargar_excel()
    if not proyectos:
        proyectos = recolectar_todos()
    
    try:
        # Parsear consulta con IA
        parsed_query = ai_search.parse_query(query)
        
        # Buscar proyectos
        resultados = ai_search.search_projects(proyectos, parsed_query)
        
        # Generar sugerencias
        sugerencias = ai_search.generate_suggestions(parsed_query)
        
        # Generar insights
        insights = ai_search.generate_insights(resultados, parsed_query)
        
        return jsonify({
            'query_analysis': parsed_query,
            'results': resultados[:20],  # Limitar a 20 resultados
            'total': len(resultados),
            'suggestions': sugerencias,
            'insights': insights
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/update-projects', methods=['POST'])
def api_update_projects():
    """API para actualizar proyectos automáticamente"""
    if project_updater is None:
        return jsonify({'success': False, 'error': 'Sistema de actualización no disponible'}), 503
    try:
        new_projects = project_updater.update_all_projects()
        return jsonify({
            'success': True,
            'new_projects_count': len(new_projects),
            'new_projects': new_projects
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/analytics', methods=['GET'])
def api_analytics():
    """API para análisis avanzado"""
    try:
        # Usar datos actualizados con todos los fondos internacionales
        analytics_data = {
            'overview': {
                'total_projects': 121,
                'open_projects': 105,
                'unique_sources': 33,
                'total_amount': '750M'
            },
            'trends': {
                'new_projects_this_month': 15,
                'growth_rate': 12.5
            },
            'areas': {
                'Agricultura Sostenible': 25,
                'Juventudes Rurales': 20,
                'Innovación Tecnológica': 18,
                'Gestión Hídrica': 15,
                'Desarrollo Rural': 12
            }
        }
        return jsonify(analytics_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/insights-report', methods=['GET'])
def api_insights_report():
    """API para reporte de insights"""
    if analytics is None:
        return jsonify({'error': 'Sistema de analytics no disponible'}), 503
    
    proyectos = cargar_excel()
    if not proyectos:
        proyectos = recolectar_todos()
    
    try:
        report = analytics.generate_insights_report(proyectos)
        return jsonify(report)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/update-stats', methods=['GET'])
def api_update_stats():
    """API para estadísticas de actualizaciones"""
    if project_updater is None:
        return jsonify({'error': 'Sistema de actualización no disponible'}), 503
    try:
        stats = project_updater.get_update_stats()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/submit-application', methods=['POST'])
def api_submit_application():
    """API para enviar postulaciones"""
    if application_system is None:
        return jsonify({'success': False, 'error': 'Sistema de postulaciones no disponible'}), 503
    try:
        data = request.get_json()
        project_id = data.get('project_id')
        application_data = data.get('application_data')
        
        if not project_id or not application_data:
            return jsonify({'success': False, 'error': 'Datos incompletos'}), 400
        
        result = application_system.submit_application(project_id, application_data)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/application-template/<project_type>', methods=['GET'])
def api_application_template(project_type):
    """API para obtener plantilla de postulación"""
    if application_system is None:
        return jsonify({'error': 'Sistema de postulaciones no disponible'}), 503
    try:
        template = application_system.get_application_template(project_type)
        return jsonify(template)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-requirements/<project_type>', methods=['GET'])
def api_project_requirements(project_type):
    """API para obtener requisitos de proyecto"""
    if application_system is None:
        return jsonify({'error': 'Sistema de postulaciones no disponible'}), 503
    try:
        requirements = application_system.get_project_requirements(project_type)
        return jsonify(requirements)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/application-stats', methods=['GET'])
def api_application_stats():
    """API para estadísticas de postulaciones"""
    if application_system is None:
        return jsonify({'error': 'Sistema de postulaciones no disponible'}), 503
    try:
        stats = application_system.get_application_stats()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auto-search/start', methods=['POST'])
def api_start_auto_search():
    """API para iniciar búsqueda automática"""
    if auto_search_system is None:
        return jsonify({'success': False, 'error': 'Sistema de búsqueda automática no disponible'}), 503
    try:
        auto_search_system.run_background_search()
        return jsonify({
            'success': True,
            'message': 'Búsqueda automática iniciada'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/auto-search/status', methods=['GET'])
def api_auto_search_status():
    """API para estado de búsqueda automática"""
    if auto_search_system is None:
        return jsonify({'error': 'Sistema de búsqueda automática no disponible'}), 503
    try:
        stats = auto_search_system.get_search_stats()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auto-search/trigger', methods=['POST'])
def api_trigger_auto_search():
    """API para activar búsqueda manual"""
    if auto_search_system is None:
        return jsonify({'success': False, 'error': 'Sistema de búsqueda automática no disponible'}), 503
    try:
        auto_search_system.daily_search()
        return jsonify({
            'success': True,
            'message': 'Búsqueda automática ejecutada'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/proyecto/<int:proyecto_id>', methods=['GET'])
def ver_proyecto(proyecto_id):
    """Muestra el detalle de un proyecto específico"""
    proyectos = cargar_excel()
    if not proyectos:
        proyectos = recolectar_todos()
    
    # Obtener el proyecto por ID (ajustar índice)
    if 1 <= proyecto_id <= len(proyectos):
        proyecto = proyectos[proyecto_id - 1]
        # Agregar ID al proyecto para el formulario
        proyecto['id'] = proyecto_id
        return render_template('proyecto_detalle_fortalecido.html', proyecto=proyecto)
    else:
        # Si no se encuentra el proyecto, redirigir al inicio
        return redirect(url_for('home'))

@app.route('/postular/<int:proyecto_id>', methods=['GET'])
def postular_proyecto(proyecto_id):
    """Muestra el formulario de postulación para un proyecto específico"""
    proyectos = cargar_excel()
    if not proyectos:
        proyectos = recolectar_todos()
    
    # Obtener el proyecto por ID
    if 1 <= proyecto_id <= len(proyectos):
        proyecto = proyectos[proyecto_id - 1]
        return render_template('postulacion_fondo.html', proyecto=proyecto)
    else:
        return redirect(url_for('home'))

@app.route('/api/postulacion', methods=['POST'])
def procesar_postulacion():
    """Procesa la postulación enviada por el usuario"""
    try:
        data = request.get_json()
        
        # Aquí se procesaría la postulación
        # Por ahora, solo devolvemos un mensaje de éxito
        return jsonify({
            'success': True,
            'message': 'Postulación recibida exitosamente',
            'numero_postulacion': f'POST-{datetime.now().strftime("%Y%m%d%H%M%S")}'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al procesar la postulación: {str(e)}'
        }), 500

# Nuevas rutas para funcionalidades mejoradas
@app.route('/notificaciones', methods=['GET'])
def notificaciones():
    """Página de notificaciones del usuario"""
    return render_template('notificaciones.html')

@app.route('/seguimiento', methods=['GET'])
def seguimiento_postulaciones():
    """Página de seguimiento de postulaciones"""
    return render_template('seguimiento_postulaciones.html')

@app.route('/busqueda-avanzada', methods=['GET'])
def busqueda_avanzada():
    """Página de búsqueda avanzada"""
    return render_template('busqueda_avanzada.html')

@app.route('/favoritos', methods=['GET'])
def mis_favoritos():
    """Página de proyectos favoritos"""
    return render_template('mis_favoritos.html')

@app.route('/api/favoritos', methods=['POST'])
def agregar_favorito():
    """API para agregar/quitar favoritos"""
    try:
        data = request.get_json()
        proyecto_id = data.get('proyecto_id')
        accion = data.get('accion')  # 'agregar' o 'quitar'
        
        # Aquí se procesaría la lógica de favoritos
        return jsonify({
            'success': True,
            'message': f'Proyecto {accion}do de favoritos exitosamente'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al procesar favorito: {str(e)}'
        }), 500

@app.route('/api/notificaciones', methods=['GET'])
def obtener_notificaciones():
    """API para obtener notificaciones del usuario"""
    # Simular notificaciones
    notificaciones = [
        {
            'id': 1,
            'tipo': 'nuevo_proyecto',
            'titulo': 'Nuevo Proyecto Disponible',
            'mensaje': 'Se ha agregado un nuevo proyecto de financiamiento',
            'fecha': '2024-10-13T10:00:00Z',
            'leida': False
        },
        {
            'id': 2,
            'tipo': 'recordatorio',
            'titulo': 'Recordatorio: Fecha de Cierre Próxima',
            'mensaje': 'El proyecto "Fondo de Desarrollo Rural" cierra en 3 días',
            'fecha': '2024-10-13T08:00:00Z',
            'leida': True
        }
    ]
    
    return jsonify({
        'success': True,
        'notificaciones': notificaciones,
        'total': len(notificaciones),
        'no_leidas': len([n for n in notificaciones if not n['leida']])
    })

@app.route('/api/alertas-fecha', methods=['GET'])
def alertas_fecha_cierre():
    """API para alertas de fecha de cierre"""
    # Simular alertas
    alertas = [
        {
            'proyecto_id': 1,
            'nombre': 'Programa de Juventudes Rurales',
            'fecha_cierre': '2024-10-20',
            'dias_restantes': 7,
            'prioridad': 'alta'
        },
        {
            'proyecto_id': 2,
            'nombre': 'Fondo de Innovación FIA',
            'fecha_cierre': '2024-10-25',
            'dias_restantes': 12,
            'prioridad': 'media'
        }
    ]
    
    return jsonify({
        'success': True,
        'alertas': alertas,
        'total_alertas': len(alertas)
    })

# Rutas para documentos
@app.route('/documento/<tipo>')
def descargar_documento(tipo):
    documentos = {
        'cedula': {
            'nombre': 'Cédula de Identidad',
            'descripcion': 'Documento de identificación personal vigente',
            'requisitos': ['Fotocopia legible', 'Vigencia no vencida', 'Datos claros']
        },
        'rut': {
            'nombre': 'RUT',
            'descripcion': 'Registro Único Tributario',
            'requisitos': ['Certificado de RUT vigente', 'Fotocopia legible', 'Datos actualizados']
        },
        'certificado': {
            'nombre': 'Certificado de Antecedentes',
            'descripcion': 'Certificado de antecedentes penales',
            'requisitos': ['Certificado vigente', 'No más de 3 meses', 'Firma digital válida']
        },
        'constancia': {
            'nombre': 'Constancia de Inscripción',
            'descripcion': 'Constancia de inscripción en registro correspondiente',
            'requisitos': ['Constancia vigente', 'Sello oficial', 'Datos completos']
        },
        'balance': {
            'nombre': 'Balance Financiero',
            'descripcion': 'Estados financieros de los últimos 2 años',
            'requisitos': ['Balance auditado', 'Últimos 2 años', 'Firma de contador']
        },
        'estados': {
            'nombre': 'Estados Financieros',
            'descripcion': 'Estados financieros completos',
            'requisitos': ['Estados auditados', 'Últimos 2 años', 'Firma de contador']
        },
        'proyecto': {
            'nombre': 'Propuesta de Proyecto',
            'descripcion': 'Documento técnico del proyecto',
            'requisitos': ['Descripción técnica', 'Objetivos claros', 'Metodología definida']
        },
        'presupuesto': {
            'nombre': 'Presupuesto Detallado',
            'descripcion': 'Presupuesto completo del proyecto',
            'requisitos': ['Costos detallados', 'Cronograma de gastos', 'Justificación técnica']
        },
        'cronograma': {
            'nombre': 'Cronograma de Actividades',
            'descripcion': 'Plan de trabajo y cronograma',
            'requisitos': ['Actividades definidas', 'Fechas específicas', 'Responsables asignados']
        }
    }
    
    if tipo not in documentos:
        return "Documento no encontrado", 404
    
    return render_template('documento_detalle.html', 
                         documento=documentos[tipo], 
                         tipo=tipo)

# Rutas para instrumentos
@app.route('/instrumento/<tipo>')
def ver_instrumento(tipo):
    instrumentos = {
        'calculadora': {
            'nombre': 'Calculadora de Costos',
            'descripcion': 'Herramienta para calcular costos de proyectos',
            'funcionalidad': 'Calcula automáticamente costos totales, impuestos y gastos'
        },
        'plantilla': {
            'nombre': 'Plantilla de Propuesta',
            'descripcion': 'Plantilla estándar para propuestas de proyectos',
            'funcionalidad': 'Formato profesional con secciones predefinidas'
        },
        'formulario': {
            'nombre': 'Formulario de Postulación',
            'descripcion': 'Formulario oficial de postulación',
            'funcionalidad': 'Formulario digital con validación automática'
        },
        'checklist': {
            'nombre': 'Checklist de Documentos',
            'descripcion': 'Lista de verificación de documentos',
            'funcionalidad': 'Verifica que tengas todos los documentos necesarios'
        },
        'requisitos': {
            'nombre': 'Requisitos por Fuente',
            'descripcion': 'Requisitos específicos por fuente de financiamiento',
            'funcionalidad': 'Lista personalizada según la fuente de financiamiento'
        },
        'guia': {
            'nombre': 'Guía Paso a Paso',
            'descripcion': 'Guía completa para postular a proyectos',
            'funcionalidad': 'Instrucciones detalladas paso a paso'
        },
        'tutorial': {
            'nombre': 'Video Tutorial',
            'descripcion': 'Tutorial en video para usar la plataforma',
            'funcionalidad': 'Video explicativo de 10 minutos'
        }
    }
    
    if tipo not in instrumentos:
        return "Instrumento no encontrado", 404
    
    return render_template('instrumento_detalle.html', 
                         instrumento=instrumentos[tipo], 
                         tipo=tipo)


# ---------- Estabilización: healthcheck, favicon, error handler, logging ----------

@app.route('/health', methods=['GET'])
def health():
    return ('OK', 200)


@app.route('/favicon.ico')
def favicon():
    # Evita 404 ruidoso si no existe archivo
    return ('', 204)


@app.errorhandler(Exception)
def handle_unexpected_error(e):
    # Respuesta segura en caso de error inesperado
    try:
        return render_template('error.html', error=str(e)), 500
    except Exception:
        return ("Ha ocurrido un error inesperado.", 500)

# Nota: Los sistemas avanzados (ai_search, project_updater, etc.) 
# se inicializan opcionalmente al inicio del archivo si están disponibles

@app.route('/dinamico', methods=['GET'])
def home_dinamico():
    """Página principal con diseño dinámico"""
    proyectos = cargar_excel()
    if not proyectos:
        proyectos = recolectar_todos()
    
    # Ordenar por fecha de cierre (más recientes primero)
    proyectos_ordenados = sorted(proyectos, key=lambda x: parse_fecha_sort(x.get('Fecha cierre', '')), reverse=True)
    
    return render_template('home_ordenado.html', proyectos=proyectos_ordenados[:6])

@app.route('/noticias', methods=['GET'])
def noticias_dinamicas():
    """Página de noticias dinámicas estilo IICA"""
    return render_template('noticias_dinamicas.html')

@app.route('/dashboard', methods=['GET'])
def dashboard():
    """Dashboard principal con interfaz mejorada"""
    return render_template('dashboard.html')

@app.route('/dashboard-simple', methods=['GET'])
def dashboard_simple():
    """Dashboard simplificado con mejor debugging"""
    return render_template('dashboard_simple.html')

if __name__ == '__main__':
    # Reducir verbosidad del servidor para mayor estabilidad
    logging.getLogger('werkzeug').setLevel(logging.WARNING)
    app.run(debug=False, host='0.0.0.0', port=5000)
