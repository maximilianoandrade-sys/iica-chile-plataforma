"""
Utilidades para trabajar con Excel sin pandas
Usa openpyxl directamente
"""

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
from typing import List, Dict, Any, Optional


def leer_excel(archivo: str) -> List[Dict[str, Any]]:
    """
    Lee un archivo Excel y retorna una lista de diccionarios.
    Reemplaza pd.read_excel() y df.to_dict('records')
    """
    if not os.path.exists(archivo):
        return []
    
    try:
        wb = load_workbook(archivo, data_only=True)
        ws = wb.active
        
        # Leer encabezados de la primera fila
        headers = []
        for cell in ws[1]:
            headers.append(cell.value if cell.value else f"Columna_{cell.column}")
        
        # Leer datos
        proyectos = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            proyecto = {}
            for idx, cell in enumerate(row):
                if idx < len(headers):
                    valor = cell.value
                    # Convertir None a string vacío para consistencia
                    proyecto[headers[idx]] = valor if valor is not None else ""
            # Solo agregar si tiene al menos un valor no vacío
            if any(v for v in proyecto.values() if v):
                proyectos.append(proyecto)
        
        return proyectos
    except Exception as e:
        print(f"❌ Error leyendo Excel {archivo}: {e}")
        return []


def guardar_excel(proyectos: List[Dict[str, Any]], archivo: str, sheet_name: str = "Proyectos"):
    """
    Guarda una lista de diccionarios en un archivo Excel.
    Reemplaza pd.DataFrame().to_excel()
    """
    if not proyectos:
        return
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Obtener todas las claves únicas de todos los proyectos
        all_keys = set()
        for proyecto in proyectos:
            all_keys.update(proyecto.keys())
        all_keys = sorted(list(all_keys))
        
        # Escribir encabezados
        for idx, key in enumerate(all_keys, start=1):
            ws.cell(row=1, column=idx, value=key)
        
        # Escribir datos
        for row_idx, proyecto in enumerate(proyectos, start=2):
            for col_idx, key in enumerate(all_keys, start=1):
                valor = proyecto.get(key, "")
                ws.cell(row=row_idx, column=col_idx, value=valor)
        
        # Asegurar que el directorio existe
        os.makedirs(os.path.dirname(archivo) if os.path.dirname(archivo) else '.', exist_ok=True)
        
        wb.save(archivo)
        print(f"💾 Guardados {len(proyectos)} proyectos en {archivo}")
    except Exception as e:
        print(f"❌ Error guardando Excel {archivo}: {e}")


def crear_dataframe_memoria(proyectos: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Función helper que simula pd.DataFrame() pero solo retorna la lista.
    Útil para compatibilidad con código existente.
    """
    return proyectos


def leer_excel_a_bytes(archivo: str) -> bytes:
    """
    Lee un archivo Excel y retorna su contenido en bytes.
    Útil para send_file de Flask.
    """
    if not os.path.exists(archivo):
        return b''
    
    with open(archivo, 'rb') as f:
        return f.read()

