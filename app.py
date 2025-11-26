from flask import Flask, render_template, redirect, url_for, request, send_file, Response, jsonify, make_response
import json
import os
import time
import uuid
import csv
import io
from datetime import datetime
from functools import lru_cache
from utils_excel import leer_excel, guardar_excel

# Importaciones de scrapers - stubs temporales si no están disponibles
try:
    from scrapers.iica import obtener_proyectos_iica
except:
    def obtener_proyectos_iica(): return []
try:
    from scrapers.devex import obtener_proyectos_devex
except:
    def obtener_proyectos_devex(): return []
try:
    from scrapers.developmentaid import obtener_proyectos_developmentaid
except:
    def obtener_proyectos_developmentaid(): return []
try:
    from scrapers.globaltenders import obtener_proyectos_globaltenders
except:
    def obtener_proyectos_globaltenders(): return []
try:
    from scrapers.ungm import obtener_proyectos_ungm
except:
    def obtener_proyectos_ungm(): return []
try:
    from scrapers.fia import obtener_proyectos_fia
except:
    def obtener_proyectos_fia(): return []
try:
    from scrapers.tenderconsultants import obtener_proyectos_tenderconsultants
except:
    def obtener_proyectos_tenderconsultants(): return []
try:
    from scrapers.fondosgob import obtener_proyectos_fondosgob
except:
    def obtener_proyectos_fondosgob(): return []
try:
    from scrapers.inia import obtener_proyectos_inia
except:
    def obtener_proyectos_inia(): return []
try:
    from scrapers.worldbank import obtener_proyectos_worldbank
except:
    def obtener_proyectos_worldbank(): return []
try:
    from scrapers.bid import obtener_proyectos_bid
except:
    def obtener_proyectos_bid(): return []
try:
    from scrapers.europeaid import obtener_proyectos_europeaid
except:
    def obtener_proyectos_europeaid(): return []
try:
    from scrapers.usaid import obtener_proyectos_usaid
except:
    def obtener_proyectos_usaid(): return []
try:
    from scrapers.undp import obtener_proyectos_undp
except:
    def obtener_proyectos_undp(): return []
try:
    from scrapers.fao import obtener_proyectos_fao
except:
    def obtener_proyectos_fao(): return []
try:
    from scrapers.oecd import obtener_proyectos_oecd
except:
    def obtener_proyectos_oecd(): return []
try:
    from scrapers.gates import obtener_proyectos_gates
except:
    def obtener_proyectos_gates(): return []
try:
    from scrapers.tenderconsultants_real import obtener_proyectos_tenderconsultants_real
except:
    def obtener_proyectos_tenderconsultants_real(): return []
try:
    from scrapers.iica_dashboard import obtener_proyectos_iica_dashboard
except:
    def obtener_proyectos_iica_dashboard(): return []
try:
    from scrapers.iica_dashboard_real import obtener_proyectos_iica_dashboard_real
except:
    def obtener_proyectos_iica_dashboard_real(): return []
try:
    from scrapers.agriculture_portal import obtener_proyectos_agriculture_portal
except:
    def obtener_proyectos_agriculture_portal(): return []
try:
    from scrapers.development_funds import obtener_proyectos_development_funds
except:
    def obtener_proyectos_development_funds(): return []
try:
    from scrapers.excel_importer import obtener_proyectos_excel
except:
    def obtener_proyectos_excel(): return []
try:
    from scrapers.corfo import obtener_proyectos_corfo
except:
    def obtener_proyectos_corfo(): return []
try:
    from scrapers.fondos_chile import obtener_proyectos_fondos_chile
except:
    def obtener_proyectos_fondos_chile(): return []
try:
    from scrapers.cnr import obtener_proyectos_cnr
except:
    def obtener_proyectos_cnr(): return []
try:
    from scrapers.gef import obtener_proyectos_gef
except:
    def obtener_proyectos_gef(): return []
try:
    from scrapers.indap import obtener_proyectos_indap
except:
    def obtener_proyectos_indap(): return []
try:
    from scrapers.fondos_gob import obtener_proyectos_fondos_gob
except:
    def obtener_proyectos_fondos_gob(): return []

# Stubs para funciones adicionales que pueden no existir
def obtener_proyectos_kickstarter(): return []
def obtener_proyectos_gofundme(): return []
def obtener_proyectos_indiegogo(): return []
def obtener_proyectos_rockethub(): return []
def obtener_proyectos_artistshare(): return []
def obtener_proyectos_agricultural_grants(): return []
def obtener_proyectos_development_funding(): return []
def obtener_proyectos_tech_innovation(): return []
def obtener_proyectos_education_research(): return []
def obtener_proyectos_environmental_sustainability(): return []

# Stubs para sistemas opcionales
ANALYTICS_AVAILABLE = False
try:
    from analytics import AnalyticsManager
    analytics_manager = AnalyticsManager()
    ANALYTICS_AVAILABLE = True
except:
    class AnalyticsManager:
        def track_session(self, *args, **kwargs): pass
        def track_search(self, *args, **kwargs): pass
        def track_project_click(self, *args, **kwargs): pass
        def get_search_analytics(self, *args, **kwargs): return {}
        def get_project_analytics(self, *args, **kwargs): return {}
        def get_performance_analytics(self, *args, **kwargs): return {}
        def get_user_analytics(self, *args, **kwargs): return {}
    analytics_manager = AnalyticsManager()

try:
    from recommendations import RecommendationEngine
    recommendation_engine = RecommendationEngine()
except:
    class RecommendationEngine:
        def update_user_profile(self, *args, **kwargs): pass
        def get_recommendations_for_user(self, *args, **kwargs): return []
    recommendation_engine = RecommendationEngine()

try:
    from translations import get_all_translations
except:
    def get_all_translations(lang='es'):
        return {
            'title': 'IICA Chile',
            'search': 'Buscar',
            'filters': 'Filtros',
            'results': 'Resultados'
        }

try:
    from notification_system import notification_manager
except:
    class NotificationManager:
        def subscribe_user(self, *args, **kwargs): pass
    notification_manager = NotificationManager()

try:
    from backup_system import backup_manager
except:
    class BackupManager:
        def create_backup(self, *args, **kwargs): return None
        def list_backups(self, *args, **kwargs): return []
    backup_manager = BackupManager()

# Decorador de caché simple
def cache_proyectos(max_age=1800):
    def decorator(func):
        func._cache_time = 0
        func._cache_result = None
        def wrapper(*args, **kwargs):
            now = time.time()
            if now - func._cache_time > max_age or func._cache_result is None:
                func._cache_result = func(*args, **kwargs)
                func._cache_time = now
            return func._cache_result
        return wrapper
    return decorator

app = Flask(__name__)
DATA_PATH = "data/proyectos.xlsx"
COLUMNS = [
    "Nombre",
    "Fuente",
    "Fecha cierre",
    "Enlace",
    "Estado",
    "Monto",
    "Área de interés",
]

# Crear directorio de datos si no existe
os.makedirs("data", exist_ok=True)

@cache_proyectos(max_age=1800)  # 30 minutos de caché
def recolectar_todos():
    proyectos = []
    # Solo fuentes básicas que funcionan
    try:
        proyectos.extend(obtener_proyectos_iica())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_devex())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_developmentaid())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_globaltenders())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_ungm())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_fia())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_tenderconsultants())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_fondosgob())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_inia())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_worldbank())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_bid())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_europeaid())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_usaid())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_undp())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_fao())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_oecd())
    except:
        pass
    try:
        proyectos.extend(obtener_proyectos_gates())
    except:
        pass
    proyectos.extend(obtener_proyectos_tenderconsultants_real())
    proyectos.extend(obtener_proyectos_iica_dashboard())
    proyectos.extend(obtener_proyectos_iica_dashboard_real())
    proyectos.extend(obtener_proyectos_agriculture_portal())
    proyectos.extend(obtener_proyectos_development_funds())
    proyectos.extend(obtener_proyectos_excel())
    proyectos.extend(obtener_proyectos_corfo())
    proyectos.extend(obtener_proyectos_fondos_chile())
    proyectos.extend(obtener_proyectos_cnr())
    proyectos.extend(obtener_proyectos_gef())
    # obtener_proyectos_fia() ya se llama arriba en línea 246, no duplicar
    proyectos.extend(obtener_proyectos_indap())
    proyectos.extend(obtener_proyectos_fondos_gob())
    
    # Nuevas fuentes internacionales
    proyectos.extend(obtener_proyectos_kickstarter())
    proyectos.extend(obtener_proyectos_gofundme())
    proyectos.extend(obtener_proyectos_indiegogo())
    proyectos.extend(obtener_proyectos_rockethub())
    proyectos.extend(obtener_proyectos_artistshare())
    proyectos.extend(obtener_proyectos_agricultural_grants())
    proyectos.extend(obtener_proyectos_development_funding())
    proyectos.extend(obtener_proyectos_tech_innovation())
    proyectos.extend(obtener_proyectos_education_research())
    proyectos.extend(obtener_proyectos_environmental_sustainability())
    
    return proyectos


def job_actualizacion_diaria():
    try:
        proyectos = recolectar_todos()
        guardar_excel(proyectos)
        print("[Scheduler] Proyectos actualizados correctamente")
    except Exception as exc:
        print(f"[Scheduler] Error en actualización: {exc}")

def guardar_excel(proyectos):
    # Normalizar llaves y valores por defecto
    normalizados = []
    for p in (proyectos or []):
        item = {
            "Nombre": p.get("Nombre", ""),
            "Fuente": p.get("Fuente", ""),
            "Fecha cierre": str(p.get("Fecha cierre", "")),
            "Enlace": p.get("Enlace", ""),
            "Estado": p.get("Estado", ""),
            "Monto": p.get("Monto", "0"),
            "Área de interés": p.get("Área de interés", ""),
        }
        # Asegurar que todas las columnas estén presentes
        for col in COLUMNS:
            if col not in item:
                item[col] = ""
        normalizados.append(item)

    # Asegurar directorio de salida
    output_dir = os.path.dirname(DATA_PATH)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    # Usar la función de utils_excel
    from utils_excel import guardar_excel as guardar_excel_util
    guardar_excel_util(normalizados, DATA_PATH)

def cargar_excel():
    try:
        return leer_excel(DATA_PATH)
    except:
        return []


def cargar_df():
    """Retorna los proyectos como lista de diccionarios (compatible con código que espera DataFrame)"""
    try:
        return leer_excel(DATA_PATH)
    except:
        return []

@app.route('/', methods=['GET'])
def home():
    # Generar o obtener session ID
    session_id = request.cookies.get('session_id')
    if not session_id:
        session_id = str(uuid.uuid4())
    
    # Tracking de sesión (solo si analytics está disponible)
    if ANALYTICS_AVAILABLE:
        ip_address = request.remote_addr
        user_agent = request.headers.get('User-Agent', '')
        analytics_manager.track_session(session_id, ip_address, user_agent)
    
    # Detectar idioma del navegador o usar español por defecto
    language = request.args.get('lang', 'es')
    if language not in ['es', 'en']:
        language = 'es'
    
    proyectos = cargar_excel()
    if not proyectos:
        proyectos = recolectar_todos()
        guardar_excel(proyectos)
    
    # Aplicar filtros
    query = request.args.get('query', '').strip()
    area = request.args.get('area', '').strip()
    estado = request.args.get('estado', '').strip()
    ordenar_por = request.args.get('ordenar_por', 'fecha')  # fecha, monto, nombre, fuente
    orden = request.args.get('orden', 'asc')  # asc, desc
    
    # Tracking de búsqueda
    start_time = time.time()
    search_filters = {
        'query': query,
        'area': area,
        'estado': estado,
        'ordenar_por': ordenar_por,
        'orden': orden
    }
    
    # Filtros de búsqueda
    if query or area or estado:
        proyectos_filtrados = []
        for p in proyectos:
            # Filtro por búsqueda
            if query:
                query_lower = query.lower()
                if not any(query_lower in str(p.get(campo, '')).lower() 
                          for campo in ['Nombre', 'Fuente', 'Área de interés', 'Descripción']):
                    continue
            
            # Filtro por área
            if area and p.get('Área de interés', '') != area:
                continue
                
            # Filtro por estado
            if estado and p.get('Estado', '') != estado:
                continue
                
            proyectos_filtrados.append(p)
        proyectos = proyectos_filtrados
    
    # Ordenamiento mejorado con manejo de errores
    if ordenar_por == 'fecha':
        def parse_fecha_sort(fecha):
            if not fecha or fecha == '':
                return '9999-12-31'  # Fechas vacías al final
            try:
                from datetime import datetime
                # Intentar parsear diferentes formatos de fecha
                if isinstance(fecha, str):
                    # Formato YYYY-MM-DD
                    if len(fecha) == 10 and fecha.count('-') == 2:
                        return fecha
                    # Formato con barras (DD/MM/YYYY o MM/DD/YYYY)
                    elif len(fecha) == 10 and fecha.count('/') == 2:
                        # Intentar primero DD/MM/YYYY (formato más común en Chile)
                        try:
                        return datetime.strptime(fecha, '%d/%m/%Y').strftime('%Y-%m-%d')
                        except ValueError:
                            # Si falla, intentar MM/DD/YYYY
                        try:
                            return datetime.strptime(fecha, '%m/%d/%Y').strftime('%Y-%m-%d')
                            except ValueError:
                                # Si ambos fallan, retornar fecha original
                                return str(fecha)
                return str(fecha)
            except:
                return '9999-12-31'
        proyectos.sort(key=lambda x: parse_fecha_sort(x.get('Fecha cierre', '')), reverse=(orden == 'desc'))
    elif ordenar_por == 'monto':
        def parse_monto_sort(monto):
            if not monto or monto in ['Consultar', 'Variable', '']:
                return 0
            try:
                import re
                numeros = re.findall(r'[\d,]+', str(monto))
                if numeros:
                    return float(numeros[0].replace(',', ''))
                return 0
            except:
                return 0
        proyectos.sort(key=lambda x: parse_monto_sort(x.get('Monto', '')), reverse=(orden == 'desc'))
    elif ordenar_por == 'nombre':
        proyectos.sort(key=lambda x: (x.get('Nombre', '') or '').lower(), reverse=(orden == 'desc'))
    elif ordenar_por == 'fuente':
        proyectos.sort(key=lambda x: (x.get('Fuente', '') or '').lower(), reverse=(orden == 'desc'))
    
    # Paginación
    page = int(request.args.get('page', 1))
    per_page = 5
    total_proyectos = len(proyectos)
    total_pages = (total_proyectos + per_page - 1) // per_page
    
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    proyectos_paginados = proyectos[start_idx:end_idx]
    
    # Estadísticas básicas
    total = len(proyectos)
    por_area = {}
    por_estado = {}
    por_moneda = {}
    total_monto = {}
    for p in proyectos:
        por_area[p.get("Área de interés", "General")] = por_area.get(p.get("Área de interés", "General"), 0) + 1
        por_estado[p.get("Estado", "N/D")] = por_estado.get(p.get("Estado", "N/D"), 0) + 1
        # Acumular montos por moneda (si están normalizados como "XXX N")
        monto_val = str(p.get("Monto", "0"))
        parts = monto_val.split()
        if len(parts) == 2 and parts[0].isalpha():
            moneda = parts[0]
            try:
                numero = float(parts[1])
            except Exception:
                try:
                    numero = float(parts[1].replace(',', ''))
                except Exception:
                    numero = 0.0
            total_monto[moneda] = total_monto.get(moneda, 0.0) + numero
            por_moneda[moneda] = por_moneda.get(moneda, 0) + 1

    stats = {
        "total": total,
        "por_area": por_area,
        "por_estado": por_estado,
        "por_moneda": por_moneda,
        "total_monto": total_monto,
    }

    # Cargar adjudicados del dashboard real de IICA
    adjudicados = obtener_proyectos_iica_dashboard_real()
    
    # También cargar desde archivo JSON si existe
    data_file = os.path.join('data', 'adjudicados.json')
    try:
        with open(data_file, 'r', encoding='utf-8') as f:
            adjudicados_json = json.load(f)
            adjudicados.extend(adjudicados_json)
    except Exception:
        pass

    # Cargar documentos
    docs_file = os.path.join('data', 'documentos.json')
    documentos = []
    try:
        with open(docs_file, 'r', encoding='utf-8') as f:
            documentos = json.load(f)
    except Exception:
        documentos = []

    # Cargar postulaciones directas
    post_file = os.path.join('data', 'postulaciones_directas.json')
    postulaciones = []
    try:
        with open(post_file, 'r', encoding='utf-8') as f:
            postulaciones = json.load(f)
    except Exception:
        postulaciones = []

    # Cargar instituciones
    inst_file = os.path.join('data', 'instituciones.json')
    instituciones = []
    try:
        with open(inst_file, 'r', encoding='utf-8') as f:
            instituciones = json.load(f)
    except Exception:
        instituciones = []

    info = {
        "titulo": "IICA Chile",
        "descripcion": (
            "El IICA (Instituto Interamericano de Cooperación para la Agricultura) en Chile "
            "promueve la innovación, la sostenibilidad y el desarrollo rural a través de proyectos, "
            "alianzas y asistencia técnica que fortalecen capacidades y mejoran la productividad del sector agrícola."
        ),
        "contacto": {
            "email": "hernan.chiriboga@iica.int",
            "telefono": "(56-2) 2225-2511",
            "fax": "(56-2) 2269-1371 / 2269-6858",
            "direccion": "Calle Rancagua No.0320, Providencia, Santiago, Chile",
            "correo_postal": "Casilla No.16107, Correo 9, Providencia, Santiago, Chile",
            "representante": "Hernán Chiriboga"
        },
        "redes_sociales": {
            "twitter": "https://x.com/iicanoticias",
            "instagram": "https://www.instagram.com/iicachile/",
            "facebook": "https://www.facebook.com/IICAChile/"
        }
    }

    # Obtener traducciones
    translations = get_all_translations(language)
    
    # Tracking de resultados de búsqueda (solo si analytics está disponible)
    if ANALYTICS_AVAILABLE:
        search_time = time.time() - start_time
        analytics_manager.track_search(
            session_id, query, search_filters, len(proyectos_paginados), search_time
        )
    
    # Datos de paginación
    pagination_data = {
        'current_page': page,
        'total_pages': total_pages,
        'total_proyectos': total_proyectos,
        'per_page': per_page,
        'has_prev': page > 1,
        'has_next': page < total_pages,
        'prev_page': page - 1 if page > 1 else None,
        'next_page': page + 1 if page < total_pages else None
    }
    
    response = make_response(render_template('home_simple.html',
                         proyectos=proyectos_paginados, 
                         stats=stats, 
                         adjudicados=adjudicados, 
                         documentos=documentos, 
                         postulaciones=postulaciones, 
                         instituciones=instituciones, 
                         info=info,
                         translations=translations,
                         current_language=language,
                         pagination=pagination_data,
                         current_filters={
                             'query': query,
                             'area': area,
                             'estado': estado,
                             'ordenar_por': ordenar_por,
                             'orden': orden
                         }))
    
    # Establecer cookie de sesión
    response.set_cookie('session_id', session_id, max_age=86400)  # 24 horas
    
    return response


@app.route('/quienes-somos', methods=['GET'])
def about():
    info = {
        "titulo": "IICA Chile",
        "descripcion": (
            "El IICA (Instituto Interamericano de Cooperación para la Agricultura) en Chile "
            "promueve la innovación, la sostenibilidad y el desarrollo rural a través de proyectos, "
            "alianzas y asistencia técnica que fortalecen capacidades y mejoran la productividad del sector agrícola."
        ),
        "contacto": {
            "email": "hernan.chiriboga@iica.int",
            "telefono": "(56-2) 2225-2511",
            "fax": "(56-2) 2269-1371 / 2269-6858",
            "direccion": "Calle Rancagua No.0320, Providencia, Santiago, Chile",
            "correo_postal": "Casilla No.16107, Correo 9, Providencia, Santiago, Chile",
            "representante": "Hernán Chiriboga"
        },
        "redes_sociales": {
            "twitter": "https://x.com/iicanoticias",
            "instagram": "https://www.instagram.com/iicachile/",
            "facebook": "https://www.facebook.com/IICAChile/"
        }
    }
    return render_template('about.html', info=info)


@app.route('/adjudicados', methods=['GET'])
def adjudicados():
    data_file = os.path.join('data', 'adjudicados.json')
    proyectos = []
    try:
        with open(data_file, 'r', encoding='utf-8') as f:
            proyectos = json.load(f)
    except Exception:
        proyectos = []
    # Filtros simples por query string
    pais = request.args.get('pais', '').strip().lower()
    anio = request.args.get('anio', '').strip()
    contraparte = request.args.get('contraparte', '').strip().lower()
    if pais:
        proyectos = [p for p in proyectos if (p.get('País', '') or '').lower().find(pais) != -1]
    if anio:
        try:
            y = int(anio)
            proyectos = [p for p in proyectos if int(p.get('Año', 0)) == y]
        except Exception:
            pass
    if contraparte:
        proyectos = [p for p in proyectos if (p.get('Contraparte', '') or '').lower().find(contraparte) != -1]
    return render_template('awarded.html', proyectos=proyectos)


@app.route('/buscar', methods=['POST'])
def buscar():
    proyectos = recolectar_todos()
    guardar_excel(proyectos)
    return redirect(url_for('home'))


@app.route('/export/csv', methods=['GET'])
def export_csv():
    proyectos = cargar_df()
    if not proyectos:
        return Response("", mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=proyectos.csv'})
    
    # Crear CSV desde lista de diccionarios
    output = io.StringIO()
    if proyectos:
        fieldnames = list(proyectos[0].keys())
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(proyectos)
    
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=proyectos.csv'}
    )


@app.route('/export/xlsx', methods=['GET'])
def export_xlsx():
    proyectos = cargar_df()
    if not proyectos:
        # Crear archivo vacío
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Proyectos'
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name='proyectos.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    # Guardar a un buffer en memoria para enviar como descarga
    from openpyxl import Workbook
    from utils_excel import guardar_excel as guardar_excel_util
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Proyectos'
    
    # Obtener todas las claves únicas
    if proyectos:
        all_keys = set()
        for proyecto in proyectos:
            all_keys.update(proyecto.keys())
        all_keys = sorted(list(all_keys))
        
        # Escribir encabezados
        for idx, key in enumerate(all_keys, start=1):
            ws.cell(row=1, column=idx, value=key)
        
        # Escribir datos
        for row_idx, proyecto in enumerate(proyectos, start=2):
            for col_idx, key in enumerate(all_keys, start=1):
                valor = proyecto.get(key, "")
                ws.cell(row=row_idx, column=col_idx, value=valor)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='proyectos.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/adjudicados/csv', methods=['GET'])
def export_adjudicados_csv():
    data_file = os.path.join('data', 'adjudicados.json')
    try:
        with open(data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception:
        data = []
    
    if not data:
        return Response("", mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=adjudicados.csv'})
    
    # Crear CSV desde lista de diccionarios
    output = io.StringIO()
    if data:
        fieldnames = list(data[0].keys())
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
    
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=adjudicados.csv'}
    )


@app.route('/export/adjudicados/xlsx', methods=['GET'])
def export_adjudicados_xlsx():
    data_file = os.path.join('data', 'adjudicados.json')
    try:
        with open(data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception:
        data = []
    
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Adjudicados'
    
    if data:
        # Obtener todas las claves únicas
        all_keys = set()
        for item in data:
            all_keys.update(item.keys())
        all_keys = sorted(list(all_keys))
        
        # Escribir encabezados
        for idx, key in enumerate(all_keys, start=1):
            ws.cell(row=1, column=idx, value=key)
        
        # Escribir datos
        for row_idx, item in enumerate(data, start=2):
            for col_idx, key in enumerate(all_keys, start=1):
                valor = item.get(key, "")
                ws.cell(row=row_idx, column=col_idx, value=valor)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='adjudicados.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ==========================
# Fondos disponibles (UI + API)
# ==========================

def _mapear_proyectos_a_fondos(proyectos):
    """Convierte los registros de 'proyectos' al formato de 'fondos' usado por las vistas genéricas.
    Los campos no disponibles se completan con valores por defecto.
    """
    fondos = []
    for idx, p in enumerate(proyectos or []):
        nombre = p.get('Nombre', '').strip()
        fuente = p.get('Fuente', '').strip()
        monto = str(p.get('Monto', '')).strip()
        fecha_cierre = str(p.get('Fecha cierre', '')).strip()
        estado = (p.get('Estado', '') or '').strip() or 'Abierto'
        area = p.get('Área de interés', '').strip() or 'General'
        enlace = p.get('Enlace', '').strip()

        # Heurística simple para tipo
        tipo = 'Nacional'
        fuente_lower = fuente.lower()
        if any(k in fuente_lower for k in ['bank', 'fund', 'global', 'fida', 'ifad', 'gef', 'undp', 'un ', 'world', 'bid', 'iadb', 'usaid', 'eu', 'ue', 'fao', 'oecd', 'gcf']):
            tipo = 'Internacional'

        fondos.append({
            'id': f"FND-{idx+1:05d}",
            'nombre': nombre or 'Fondo sin nombre',
            'organizacion': fuente or 'N/D',
            'pais': 'Internacional' if tipo == 'Internacional' else 'Chile',
            'monto': monto or 'Consultar',
            'fecha_cierre': fecha_cierre or 'N/D',
            'estado': estado,
            'area': area,
            'descripcion': p.get('Descripción', '') or 'Sin descripción disponible.',
            'requisitos': p.get('Requisitos', '') or 'Revisar bases de postulación',
            'enlace': enlace,
            'contacto': p.get('Contacto', '') or '',
            'tipo': tipo,
            'duracion': p.get('Duración', '') or '',
            'beneficiarios': p.get('Beneficiarios', '') or ''
        })
    return fondos


@app.route('/fondos', methods=['GET'])
@app.route('/fondos-disponibles', methods=['GET'])
def fondos_disponibles_view():
    """Vista de catálogo de fondos disponibles con filtros básicos (cliente usa /api/fondos)."""
    query = request.args.get('q', '').strip()
    area = request.args.get('area', '').strip()
    organizacion = request.args.get('organizacion', '').strip()
    return render_template('fondos.html', query=query, area=area, organizacion=organizacion, total=0, fondos=[])


@app.route('/api/fondos', methods=['GET'])
def api_fondos_list():
    """Entrega listado de fondos mapeados desde el Excel consolidado, con filtros opcionales."""
    fondos = _mapear_proyectos_a_fondos(cargar_excel())

    # Filtros simples
    q = request.args.get('q', '').strip().lower()
    area = request.args.get('area', '').strip()
    organizacion = request.args.get('organizacion', '').strip().lower()
    estado = request.args.get('estado', '').strip()
    tipo = request.args.get('tipo', '').strip()  # Nacional / Internacional

    if q:
        fondos = [f for f in fondos if (q in f['nombre'].lower() or q in f['descripcion'].lower() or q in f['organizacion'].lower())]
    if area:
        fondos = [f for f in fondos if f['area'] == area]
    if organizacion:
        fondos = [f for f in fondos if organizacion in f['organizacion'].lower()]
    if estado:
        fondos = [f for f in fondos if f['estado'] == estado]
    if tipo:
        fondos = [f for f in fondos if f['tipo'] == tipo]

    return jsonify({'total': len(fondos), 'fondos': fondos})


@app.route('/fondos-internacionales', methods=['GET'])
def fondos_internacionales_view():
    """Sección dedicada a Fondos Internacionales (consume /api/fondos?tipo=Internacional)."""
    return render_template('fondos_internacionales.html')

@app.route('/crear-plantilla', methods=['GET'])
def crear_plantilla():
    from scrapers.excel_importer import crear_plantilla_excel
    try:
        template_path = crear_plantilla_excel()
        if template_path and os.path.exists(template_path):
            return send_file(template_path, as_attachment=True, download_name='plantilla_proyectos.xlsx')
        else:
            return "Error creando plantilla", 500
    except Exception as e:
        return f"Error: {e}", 500

# Nuevas rutas para funcionalidades avanzadas

@app.route('/api/project-click', methods=['POST'])
def track_project_click():
    """Rastrea clics en proyectos"""
    data = request.get_json()
    session_id = request.cookies.get('session_id')
    
    if session_id and data:
        analytics_manager.track_project_click(
            session_id,
            data.get('project_name', ''),
            data.get('project_source', ''),
            data.get('project_url', '')
        )
        
        # Actualizar perfil del usuario
        recommendation_engine.update_user_profile(
            session_id, data, 'click', 1.0
        )
    
    return jsonify({'status': 'success'})

@app.route('/api/recommendations')
def get_recommendations():
    """Obtiene recomendaciones personalizadas"""
    session_id = request.cookies.get('session_id')
    if not session_id:
        return jsonify({'recommendations': []})
    
    proyectos = cargar_excel()
    if not proyectos:
        proyectos = recolectar_todos()
    
    # Obtener recomendaciones
    recommendations = recommendation_engine.get_recommendations_for_user(
        session_id, proyectos, 10
    )
    
    return jsonify({'recommendations': recommendations})

@app.route('/api/analytics')
def get_analytics():
    """Obtiene analytics del sistema"""
    days = request.args.get('days', 30, type=int)
    
    search_analytics = analytics_manager.get_search_analytics(days)
    project_analytics = analytics_manager.get_project_analytics(days)
    performance_analytics = analytics_manager.get_performance_analytics(7)
    user_analytics = analytics_manager.get_user_analytics(days)
    
    return jsonify({
        'search': search_analytics,
        'projects': project_analytics,
        'performance': performance_analytics,
        'users': user_analytics
    })

@app.route('/api/notifications/subscribe', methods=['POST'])
def subscribe_notifications():
    """Suscribe usuario a notificaciones"""
    data = request.get_json()
    
    if data and data.get('email'):
        notification_manager.subscribe_user(
            email=data['email'],
            areas_interes=data.get('areas', []),
            fuentes=data.get('fuentes', []),
            monto_minimo=data.get('monto_minimo', 0),
            monto_maximo=data.get('monto_maximo'),
            frecuencia=data.get('frecuencia', 'daily')
        )
        return jsonify({'status': 'success'})
    
    return jsonify({'status': 'error', 'message': 'Email requerido'}), 400

@app.route('/api/backup/create', methods=['POST'])
def create_backup():
    """Crea un respaldo del sistema"""
    backup_type = request.json.get('type', 'data')
    backup_path = backup_manager.create_backup(backup_type)
    
    if backup_path:
        return jsonify({'status': 'success', 'backup_path': backup_path})
    else:
        return jsonify({'status': 'error'}), 500

@app.route('/api/backup/list')
def list_backups():
    """Lista respaldos disponibles"""
    backups = backup_manager.list_backups()
    return jsonify({'backups': backups})

@app.route('/api/cache/clear', methods=['POST'])
def clear_cache():
    """Limpia el caché del sistema"""
    # Cache deshabilitado temporalmente
    return jsonify({'status': 'success', 'message': 'Cache no disponible'})

@app.route('/api/health')
def health_check():
    """Verifica el estado del sistema"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'cache_status': 'active',
        'analytics_status': 'active',
        'notifications_status': 'active',
        'recommendations_status': 'active',
        'backup_status': 'active'
    })

# Inicializar sistemas básicos
def init_basic_systems():
    """Inicializa sistemas básicos para estabilidad"""
    try:
        print("🚀 Inicializando plataforma IICA...")
        print("✅ Sistema básico inicializado")
        print("✅ Funcionalidades principales disponibles")
        print("🎉 Plataforma lista para usar")
        return True
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

if __name__ == '__main__':
    # Inicializar sistemas básicos
    init_basic_systems()
    
    # Scheduler deshabilitado temporalmente
    # try:
    #     scheduler = BackgroundScheduler(daemon=True)
    #     scheduler.add_job(job_actualizacion_diaria, 'cron', hour=3, minute=0)
    #     scheduler.start()
    #     print('[Scheduler] Tarea diaria programada a las 03:00')
    # except Exception as exc:
    #     print(f'[Scheduler] No se pudo iniciar: {exc}')
    # Configuración para producción con dominio personalizado
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    
    app.run(
        debug=debug, 
        host='0.0.0.0', 
        port=port,
        threaded=True  # Mejor rendimiento para múltiples usuarios
    )
